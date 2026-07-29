from __future__ import annotations

def configure_databank_feature_dependencies(namespace):
    # Shared imports, constants, logging, and DB helpers remain owned by the foundation app.
    for name, value in namespace.items():
        if not str(name).startswith('__'):
            globals()[name] = value


def _clear_preview(self):
    # Preview removed; keep as no-op for backward compatibility
    return


def _get_databank_focus_date(self):
    from datetime import date as _date
    day = getattr(self, '_dbank_last_day', None)
    try:
        if day:
            return _date(int(self.grid_year), int(self.grid_month), int(day)).strftime('%Y-%m-%d')
    except Exception:
        pass
    try:
        today = _date.today()
        gy = int(getattr(self, 'grid_year', today.year))
        gm = int(getattr(self, 'grid_month', today.month))
        if gy == today.year and gm == today.month:
            return today.strftime('%Y-%m-%d')
        return _date(gy, gm, 1).strftime('%Y-%m-%d')
    except Exception:
        try:
            return _date.today().strftime('%Y-%m-%d')
        except Exception as __spina_exc:
            __spina_logger = globals().get('_log_suppressed_once')
            if callable(__spina_logger):
                __spina_logger('silent_ui_9964__get_databank_focus_date', 'suppressed UI/startup exception at line 9964', __spina_exc)
            return ''


def _show_system_data_tab(self):
    try:
        tabs = set(self.nb.tabs())
        if str(self.tab_system_data) not in tabs:
            self.nb.add(self.tab_system_data, text='Data')
        else:
            self.nb.tab(self.tab_system_data, text='Data')
    except Exception as __spina_exc:
        _log_suppressed_once('excpass_system_data_tab_show', 'suppressed exception excpass_system_data_tab_show', __spina_exc)
        pass


def _hide_system_data_tab(self):
    try:
        self.nb.hide(self.tab_system_data)
    except Exception as __spina_exc:
        _log_suppressed_once('excpass_system_data_tab_hide', 'suppressed exception excpass_system_data_tab_hide', __spina_exc)
        pass


def _system_data_open_close(self):
    ds = self._system_data_get_date()
    if not ds:
        return
    try:
        self.open_databank_close_dialog(ds)
    finally:
        try:
            self._system_data_refresh_summary()
        except Exception:
            pass


def _system_data_open_history(self):
    ds = self._system_data_get_date()
    if not ds:
        return
    self.open_databank_close_history_dialog(ds)


def _system_data_open_records(self):
    ds = self._system_data_get_date()
    if not ds:
        return
    self.open_databank_close_records_dialog(start_date=ds, end_date=ds)


def _system_data_print_report(self):
    ds = self._system_data_get_date()
    if not ds:
        return
    self.print_databank_close_report(ds)


def _load_collectors_route_map(self):
    os.makedirs(DATA_DIR, exist_ok=True)
    path = data_path("collectors.json")
    try:
        with open(path, 'r', encoding='utf-8') as _f:
            raw = json.load(_f)
    except Exception:
        raw = {}

    data = {}
    if isinstance(raw, dict):
        for name, rec in (raw or {}).items():
            if not name:
                continue
            if isinstance(rec, dict):
                areas = rec.get('areas') or rec.get('route') or []
                notes = rec.get('notes') or ''
            elif isinstance(rec, list):
                areas = rec
                notes = ''
            else:
                areas = []
                notes = ''
            areas = [str(a).strip() for a in (areas or []) if str(a).strip()]
            data[str(name).strip()] = {'areas': areas, 'notes': str(notes or '')}
    elif isinstance(raw, list):
        for el in raw:
            if isinstance(el, dict):
                name = (el.get('name') or el.get('collector') or '').strip()
                if not name:
                    continue
                areas = [str(a).strip() for a in (el.get('areas') or el.get('route') or []) if str(a).strip()]
                if name in data:
                    cur_areas = data[name]['areas']
                    for a in areas:
                        if a not in cur_areas:
                            cur_areas.append(a)
                else:
                    data[name] = {'areas': areas, 'notes': str(el.get('notes') or '')}
    return data


def _build_databank_collector_defaults_for_date(self, date_s):
    def _norm_area(s: str) -> str:
        try:
            return ' '.join(str(s or '').split()).strip().lower()
        except Exception:
            return str(s or '').strip().lower()

    routes = self._load_collectors_route_map()
    cur = self.db.conn.cursor()
    area_totals = {}
    try:
        rows = cur.execute(
            """
                SELECT COALESCE(NULLIF(TRIM(c.area),''),'(No Area)') AS area_name,
                       COALESCE(SUM(COALESCE(t.payment,0)), 0) AS total_payment
                  FROM transactions t
                  LEFT JOIN clients c
                    ON (TRIM(IFNULL(t.client_uid,'')) <> '' AND c.client_uid = t.client_uid)
                    OR ((TRIM(IFNULL(t.client_uid,'')) = '' OR t.client_uid IS NULL)
                        AND c.name = t.name
                        AND IFNULL(NULLIF(TRIM(c.loan_type),''),'Regular') = IFNULL(NULLIF(TRIM(t.loan_type),''),'Regular'))
                 WHERE date(t.date)=date(?)
                 GROUP BY COALESCE(NULLIF(TRIM(c.area),''),'(No Area)')
                """,
            (date_s,),
        ).fetchall()
        for r in (rows or []):
            try:
                area_name = r['area_name'] if hasattr(r, 'keys') else r[0]
            except Exception:
                area_name = ''
            try:
                total_payment = float((r['total_payment'] if hasattr(r, 'keys') else r[1]) or 0.0)
            except Exception:
                total_payment = 0.0
            area_totals[_norm_area(area_name)] = round(total_payment, 2)
    except Exception:
        area_totals = {}

    built = []
    covered_total = 0.0
    for idx, name in enumerate(sorted(routes.keys(), key=lambda s: s.lower())):
        rec = routes.get(name) or {}
        total = 0.0
        for area in (rec.get('areas') or []):
            total += float(area_totals.get(_norm_area(area), 0.0) or 0.0)
        total = round(total, 2)
        covered_total += total
        built.append({
            'collector_name': str(name).strip(),
            'expected_amount': total,
            'actual_cash': 0.0,
            'note': '',
            'sort_order': idx,
        })

    try:
        day_expected = round(float(self.db.get_databank_daily_total(date_s, loan_type='__ALL__') or 0.0), 2)
    except Exception:
        day_expected = 0.0
    diff = round(day_expected - covered_total, 2)
    if abs(diff) >= 0.005:
        built.append({
            'collector_name': 'Unassigned / Other',
            'expected_amount': diff,
            'actual_cash': 0.0,
            'note': 'Auto-added so split expected can match the system total.',
            'sort_order': len(built),
        })
    return built


def print_databank_close_report(self, date_s=None):
    import os
    from tkinter import filedialog, messagebox
    try:
        from reportlab.pdfgen import canvas as _cv
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors as _rl_colors
    except Exception as e:
        messagebox.showerror('Daily Close Report', "Printing requires ReportLab.\n\nInstall it with:\n  pip install reportlab")
        try:
            _log_exc('print_databank_close_report:missing_reportlab', e)
        except Exception:
            pass
        return

    ds = (date_s or self._get_databank_focus_date() or '').strip()
    if not ds:
        messagebox.showerror('Daily Close Report', 'No date was selected.')
        return

    rec = None
    try:
        rec = self.db.get_databank_day_close(ds) if hasattr(self, 'db') else None
    except Exception:
        rec = None
    if not rec:
        messagebox.showerror('Daily Close Report', 'No Daily Close record exists yet for that date.')
        return

    try:
        reg_expected = round(float(self.db.get_databank_daily_total(ds, loan_type='Regular') or 0.0), 2)
    except Exception:
        reg_expected = 0.0
    try:
        x7_expected = round(float(self.db.get_databank_daily_total(ds, loan_type='7x7') or 0.0), 2)
    except Exception:
        x7_expected = 0.0
    try:
        combined_expected = round(float(rec.get('expected_amount') or 0.0), 2)
    except Exception:
        combined_expected = round(reg_expected + x7_expected, 2)
    try:
        actual_cash = round(float(rec.get('actual_cash') or 0.0), 2)
    except Exception:
        actual_cash = 0.0
    try:
        variance = round(float(rec.get('variance') or 0.0), 2)
    except Exception:
        variance = round(actual_cash - combined_expected, 2)
    if abs(variance) < 0.005:
        variance = 0.0
    status_txt = ((rec.get('variance_status') or ('Balanced' if variance == 0 else ('Overage' if variance > 0 else 'Short'))) or '').strip() or 'Balanced'
    workflow_txt = ((rec.get('variance_workflow_status') or ('Resolved' if variance == 0 else 'Pending')) or '').strip() or 'Open'
    lock_txt = 'CLOSED' if bool(int(rec.get('is_closed') or 0)) else 'OPEN'
    note_txt = (rec.get('note') or '').strip()
    try:
        collector_rows = self.db.list_databank_day_collectors(ds) if hasattr(self, 'db') else []
    except Exception:
        collector_rows = []
    closed_by = (rec.get('closed_by') or '').strip() or '—'
    closed_at = (rec.get('closed_at') or '').strip() or '—'
    opened_by = (rec.get('opened_by') or '').strip() or '—'
    opened_at = (rec.get('opened_at') or '').strip() or '—'
    updated_at = (rec.get('updated_at') or '').strip() or '—'

    tx_count = 0
    try:
        row = self.db.conn.cursor().execute(
            "SELECT COUNT(*) FROM transactions WHERE date(date)=date(?)",
            (ds,),
        ).fetchone()
        tx_count = int((row[0] if row else 0) or 0)
    except Exception:
        tx_count = 0

    out_path = filedialog.asksaveasfilename(
        parent=self.root,
        title='Save Daily Cash Close Report',
        defaultextension='.pdf',
        initialfile=f'Daily_Cash_Close_{ds}.pdf',
        filetypes=[('PDF files', '*.pdf')],
    )
    if not out_path:
        return

    try:
        c = _cv.Canvas(out_path, pagesize=A4)
        page_w, page_h = A4
        margin = 42
        line_y = page_h - margin

        def _line(text, *, x=None, y=None, size=10, bold=False, color=None):
            nonlocal line_y
            xx = margin if x is None else x
            yy = line_y if y is None else y
            try:
                if color is not None:
                    c.setFillColor(color)
                else:
                    c.setFillColorRGB(0, 0, 0)
            except Exception:
                pass
            try:
                c.setFont('Helvetica-Bold' if bold else 'Helvetica', size)
            except Exception:
                pass
            c.drawString(xx, yy, str(text))
            if y is None:
                line_y -= max(size + 6, 14)

        def _money(v):
            return fmt_currency(v)

        def _hr(gap_before=4, gap_after=10):
            nonlocal line_y
            line_y -= gap_before
            c.setStrokeColorRGB(0.75, 0.75, 0.75)
            c.setLineWidth(0.8)
            c.line(margin, line_y, page_w - margin, line_y)
            line_y -= gap_after

        def _kv(label, value, value_bold=False):
            nonlocal line_y
            lbl_w = 120
            c.setFont('Helvetica-Bold', 10)
            c.drawString(margin, line_y, f'{label}')
            c.setFont('Helvetica-Bold' if value_bold else 'Helvetica', 10)
            c.drawString(margin + lbl_w, line_y, str(value))
            line_y -= 16

        _line('SPINA DATA BANK DAILY CASH CLOSE REPORT', size=14, bold=True)
        _line('Combined Daily Close for Regular + 7x7', size=10)
        _hr()

        _kv('Date', ds, value_bold=True)
        _kv('Lock State', lock_txt, value_bold=True)
        _kv('Workflow', workflow_txt, value_bold=True)
        _kv('Status', status_txt, value_bold=True)
        _kv('Coverage', 'Combined (Regular + 7x7)')
        _kv('Transactions', tx_count)
        _hr()

        _line('DAILY TOTALS', size=11, bold=True)
        _kv('Regular Total', _money(reg_expected))
        _kv('7x7 Total', _money(x7_expected))
        _kv('Combined Expected', _money(combined_expected), value_bold=True)
        _kv('Actual Cash', _money(actual_cash), value_bold=True)
        _kv('Variance', _money(abs(variance)) if variance else _money(0.0), value_bold=True)
        _kv('Variance Type', status_txt)
        _hr()

        _line('CLOSE RECORD', size=11, bold=True)
        _kv('Closed By', closed_by)
        _kv('Closed At', closed_at)
        _kv('Reopened By', opened_by)
        _kv('Reopened At', opened_at)
        _kv('Last Updated', updated_at)
        _hr()

        if collector_rows:
            _hr()
            _line('COLLECTOR SPLIT', size=11, bold=True)
            for crow in collector_rows:
                if line_y < margin + 72:
                    c.showPage()
                    line_y = page_h - margin
                cname = (crow.get('collector_name') or '').strip() or '—'
                cexp = _money(crow.get('expected_amount') or 0.0)
                cact = _money(crow.get('actual_cash') or 0.0)
                try:
                    cvar_raw = round(float(crow.get('variance') or 0.0), 2)
                except Exception:
                    cvar_raw = round(float(crow.get('actual_cash') or 0.0) - float(crow.get('expected_amount') or 0.0), 2)
                if abs(cvar_raw) < 0.005:
                    cvar_raw = 0.0
                cvar = _money(abs(cvar_raw)) if cvar_raw else _money(0.0)
                cnote = (crow.get('note') or '').strip()
                _kv(cname, f'Expected {cexp}   Actual {cact}   Variance {cvar}')
                if cnote:
                    _kv(' ', cnote)

        _line('NOTE', size=11, bold=True)
        note_block = note_txt or '—'
        max_width = page_w - (margin * 2)
        words = note_block.split()
        wrapped = []
        if not words:
            wrapped = ['—']
        else:
            cur = words[0]
            for w in words[1:]:
                test = cur + ' ' + w
                try:
                    width = c.stringWidth(test, 'Helvetica', 10)
                except Exception:
                    width = len(test) * 5
                if width <= max_width:
                    cur = test
                else:
                    wrapped.append(cur)
                    cur = w
            wrapped.append(cur)
        c.setFont('Helvetica', 10)
        for ln in wrapped:
            if line_y < margin + 40:
                c.showPage()
                line_y = page_h - margin
                c.setFont('Helvetica', 10)
            c.drawString(margin, line_y, ln)
            line_y -= 14

        line_y = max(line_y - 8, margin + 28)
        c.setFont('Helvetica-Oblique', 8)
        c.setFillColor(_rl_colors.grey)
        c.drawString(margin, margin - 2 + 14, 'Generated from Data Bank Daily Close / Variance')
        c.drawRightString(page_w - margin, margin - 2 + 14, f'Printed: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
        c.save()
    except Exception as e:
        try:
            if out_path and os.path.exists(out_path):
                os.remove(out_path)
        except Exception:
            pass
        try:
            _log_exc('print_databank_close_report', e)
        except Exception:
            pass
        messagebox.showerror('Daily Close Report', f'Could not generate the PDF.\n\n{e}')
        return

    try:
        _open_path(out_path)
    except Exception:
        pass
    messagebox.showinfo('Daily Close Report', f'Report saved to:\n{out_path}')


def open_databank_close_dialog(self, date_s=None):
    import tkinter as tk
    from tkinter import ttk, messagebox
    from datetime import datetime as _dt

    top = tk.Toplevel(self.root)
    top.title('Data Bank Daily Close')
    top.transient(self.root)
    top.grab_set()
    top.resizable(True, True)
    try:
        top.geometry('1080x720')
    except Exception:
        pass

    outer = ttk.Frame(top, padding=12)
    outer.pack(fill='both', expand=True)
    outer.columnconfigure(1, weight=1)
    outer.columnconfigure(2, weight=1)
    outer.rowconfigure(12, weight=1)

    ttk.Label(outer, text='Daily Close / Variance', style='Section.TLabel').grid(row=0, column=0, columnspan=4, sticky='w')
    ttk.Label(
        outer,
        text='You can view any day. Closing locks the combined day for both Regular and 7x7 until reopened with a password. Split-by-collector is optional but recommended for audit.',
    ).grid(row=1, column=0, columnspan=4, sticky='w', pady=(4, 4))

    can_edit_close = ((getattr(self, 'user_role', '') or '').strip() == 'System')
    edit_hint = 'Editing is allowed only while logged in as the System user.' if not can_edit_close else 'System user: closing, reopening, and workflow updates require your password.'
    ttk.Label(outer, text=edit_hint, foreground='#666666').grid(row=2, column=0, columnspan=4, sticky='w', pady=(0, 10))

    date_var = tk.StringVar(value=(date_s or self._get_databank_focus_date() or ''))
    mode_var = tk.StringVar(value='Combined (Regular + 7x7)')
    expected_var = tk.StringVar(value='0.00')
    split_expected_var = tk.StringVar(value='0.00')
    actual_var = tk.StringVar(value='0.00')
    split_actual_var = tk.StringVar(value='0.00')
    variance_var = tk.StringVar(value='0.00')
    status_var = tk.StringVar(value='Balanced')
    workflow_var = tk.StringVar(value='Open')
    lock_var = tk.StringVar(value='OPEN')
    meta_var = tk.StringVar(value='')
    note_var = tk.StringVar(value='')
    split_check_var = tk.StringVar(value='No collector split yet.')
    collector_name_var = tk.StringVar(value='')
    collector_expected_var = tk.StringVar(value='0.00')
    collector_actual_var = tk.StringVar(value='0.00')
    collector_note_var = tk.StringVar(value='')
    expected_box = {'value': 0.0}
    state = {'record': None, 'collector_rows': [], 'selected_index': None}

    def _parse_amount(txt):
        try:
            s = str(txt or '').replace(',', '').strip()
            return float(s) if s else 0.0
        except Exception:
            return 0.0

    def _fmt_amount(val):
        try:
            return fmt_currency(val)
        except Exception:
            try:
                return f"{float(val or 0.0):,.2f}"
            except Exception:
                return '0.00'

    def _variance_status(v):
        try:
            vv = float(v or 0.0)
        except Exception:
            vv = 0.0
        if abs(vv) < 0.005:
            return 'Balanced'
        return 'Overage' if vv > 0 else 'Short'

    def _default_workflow(v):
        try:
            vv = float(v or 0.0)
        except Exception:
            vv = 0.0
        return 'Resolved' if abs(vv) < 0.005 else 'Pending'

    ttk.Label(outer, text='Date').grid(row=3, column=0, sticky='w', pady=4)
    ttk.Entry(outer, textvariable=date_var, width=16).grid(row=3, column=1, sticky='w', pady=4)
    ttk.Label(outer, text='Coverage').grid(row=4, column=0, sticky='w', pady=4)
    ttk.Label(outer, textvariable=mode_var).grid(row=4, column=1, sticky='w', pady=4)
    ttk.Label(outer, text='System Expected').grid(row=5, column=0, sticky='w', pady=4)
    ttk.Label(outer, textvariable=expected_var).grid(row=5, column=1, sticky='w', pady=4)
    ttk.Label(outer, text='Collector Expected').grid(row=5, column=2, sticky='w', pady=4)
    ttk.Label(outer, textvariable=split_expected_var).grid(row=5, column=3, sticky='w', pady=4)
    ttk.Label(outer, text='Actual Cash').grid(row=6, column=0, sticky='w', pady=4)
    actual_ent = ttk.Entry(outer, textvariable=actual_var, width=18)
    actual_ent.grid(row=6, column=1, sticky='w', pady=4)
    ttk.Label(outer, text='Collector Actual').grid(row=6, column=2, sticky='w', pady=4)
    ttk.Label(outer, textvariable=split_actual_var).grid(row=6, column=3, sticky='w', pady=4)
    ttk.Label(outer, text='Variance').grid(row=7, column=0, sticky='w', pady=4)
    ttk.Label(outer, textvariable=variance_var).grid(row=7, column=1, sticky='w', pady=4)
    ttk.Label(outer, text='Status').grid(row=7, column=2, sticky='w', pady=4)
    ttk.Label(outer, textvariable=status_var).grid(row=7, column=3, sticky='w', pady=4)
    ttk.Label(outer, text='Workflow').grid(row=8, column=0, sticky='w', pady=4)
    workflow_cmb = ttk.Combobox(outer, textvariable=workflow_var, values=('Open', 'Pending', 'Resolved'), state='readonly', width=18)
    workflow_cmb.grid(row=8, column=1, sticky='w', pady=4)
    ttk.Label(outer, text='Lock State').grid(row=8, column=2, sticky='w', pady=4)
    ttk.Label(outer, textvariable=lock_var).grid(row=8, column=3, sticky='w', pady=4)
    ttk.Label(outer, text='Note').grid(row=9, column=0, sticky='w', pady=4)
    note_ent = ttk.Entry(outer, textvariable=note_var, width=80)
    note_ent.grid(row=9, column=1, columnspan=3, sticky='ew', pady=4)
    ttk.Label(outer, textvariable=meta_var, foreground='#666666').grid(row=10, column=0, columnspan=4, sticky='w', pady=(6, 4))
    ttk.Label(outer, textvariable=split_check_var, foreground='#666666').grid(row=11, column=0, columnspan=4, sticky='w', pady=(0, 8))

    split_box = ttk.LabelFrame(outer, text='Collector Cash Split (Optional but Recommended)')
    split_box.grid(row=12, column=0, columnspan=4, sticky='nsew', pady=(0, 8))
    split_box.columnconfigure(0, weight=1)
    split_box.rowconfigure(0, weight=1)

    split_cols = ('collector', 'expected', 'actual', 'variance', 'note')
    split_tree = ttk.Treeview(split_box, columns=split_cols, show='headings', height=10)
    split_tree.heading('collector', text='Collector')
    split_tree.heading('expected', text='Expected')
    split_tree.heading('actual', text='Actual Cash')
    split_tree.heading('variance', text='Variance')
    split_tree.heading('note', text='Note')
    split_tree.column('collector', width=190, anchor='w', stretch=False)
    split_tree.column('expected', width=110, anchor='e', stretch=False)
    split_tree.column('actual', width=110, anchor='e', stretch=False)
    split_tree.column('variance', width=110, anchor='e', stretch=False)
    split_tree.column('note', width=340, anchor='w', stretch=True)
    svsb = ttk.Scrollbar(split_box, orient='vertical', command=split_tree.yview)
    split_tree.configure(yscrollcommand=svsb.set)
    split_tree.grid(row=0, column=0, sticky='nsew', padx=(8, 0), pady=8)
    svsb.grid(row=0, column=1, sticky='ns', pady=8)

    editor = ttk.Frame(split_box)
    editor.grid(row=1, column=0, columnspan=2, sticky='ew', padx=8, pady=(0, 8))
    editor.columnconfigure(7, weight=1)
    ttk.Label(editor, text='Collector').grid(row=0, column=0, sticky='w')
    ttk.Entry(editor, textvariable=collector_name_var, width=22).grid(row=0, column=1, sticky='w', padx=(4, 10))
    ttk.Label(editor, text='Expected').grid(row=0, column=2, sticky='w')
    ttk.Entry(editor, textvariable=collector_expected_var, width=12).grid(row=0, column=3, sticky='w', padx=(4, 10))
    ttk.Label(editor, text='Actual').grid(row=0, column=4, sticky='w')
    ttk.Entry(editor, textvariable=collector_actual_var, width=12).grid(row=0, column=5, sticky='w', padx=(4, 10))
    ttk.Label(editor, text='Note').grid(row=0, column=6, sticky='w')
    ttk.Entry(editor, textvariable=collector_note_var, width=36).grid(row=0, column=7, sticky='ew', padx=(4, 10))

    editor_btns = ttk.Frame(split_box)
    editor_btns.grid(row=2, column=0, columnspan=2, sticky='e', padx=8, pady=(0, 8))
    add_update_btn = ttk.Button(editor_btns, text='Add / Update Row')
    add_update_btn.pack(side='left', padx=4)
    remove_row_btn = ttk.Button(editor_btns, text='Remove Selected')
    remove_row_btn.pack(side='left', padx=4)
    load_routes_btn = ttk.Button(editor_btns, text='Load Collectors from Routes')
    load_routes_btn.pack(side='left', padx=4)
    save_split_btn = ttk.Button(editor_btns, text='Save Split')
    save_split_btn.pack(side='left', padx=4)
    use_split_btn = ttk.Button(editor_btns, text='Use Collector Actual Total')
    use_split_btn.pack(side='left', padx=4)
    history_btn = ttk.Button(editor_btns, text='View History')
    history_btn.pack(side='left', padx=4)

    btns = ttk.Frame(outer)
    btns.grid(row=13, column=0, columnspan=4, sticky='e', pady=(4, 0))
    load_btn = ttk.Button(btns, text='Load')
    load_btn.pack(side='left', padx=4)
    close_btn = ttk.Button(btns, text='Close Day')
    close_btn.pack(side='left', padx=4)
    reopen_btn = ttk.Button(btns, text='Reopen (Password)')
    reopen_btn.pack(side='left', padx=4)
    workflow_btn = ttk.Button(btns, text='Update Workflow')
    workflow_btn.pack(side='left', padx=4)
    print_btn = ttk.Button(btns, text='Print Report')
    print_btn.pack(side='left', padx=4)
    route_copy_btn = ttk.Button(btns, text='Save Route Copy')
    route_copy_btn.pack(side='left', padx=4)
    list_btn = ttk.Button(btns, text='Close Records List')
    list_btn.pack(side='left', padx=4)
    ttk.Button(btns, text='Close Window', command=top.destroy).pack(side='left', padx=4)

    def _selected_split_index():
        sel = split_tree.selection()
        if not sel:
            return None
        try:
            return int(sel[0])
        except Exception:
            return None

    def _clear_split_editor():
        state['selected_index'] = None
        collector_name_var.set('')
        collector_expected_var.set('0.00')
        collector_actual_var.set('0.00')
        collector_note_var.set('')

    def _refresh_split_tree(select_index=None):
        for iid in split_tree.get_children():
            split_tree.delete(iid)
        rows = list(state.get('collector_rows') or [])
        split_expected = 0.0
        split_actual = 0.0
        for idx, row in enumerate(rows):
            try:
                exp = round(float(row.get('expected_amount') or 0.0), 2)
            except Exception:
                exp = 0.0
            try:
                act = round(float(row.get('actual_cash') or 0.0), 2)
            except Exception:
                act = 0.0
            var = round(act - exp, 2)
            if abs(var) < 0.005:
                var = 0.0
            split_expected += exp
            split_actual += act
            split_tree.insert('', 'end', iid=str(idx), values=(
                row.get('collector_name') or '',
                _fmt_amount(exp),
                _fmt_amount(act),
                _fmt_amount(abs(var)) if var else _fmt_amount(0.0),
                row.get('note') or '',
            ))
        split_expected = round(split_expected, 2)
        split_actual = round(split_actual, 2)
        split_expected_var.set(_fmt_amount(split_expected))
        split_actual_var.set(_fmt_amount(split_actual))
        day_expected = round(float(expected_box.get('value') or 0.0), 2)
        diff = round(split_expected - day_expected, 2)
        if rows:
            if abs(diff) < 0.005:
                split_check_var.set('Collector expected matches the system total for this day.')
            elif diff > 0:
                split_check_var.set(f'Collector expected is OVER the system total by {_fmt_amount(diff)}. Check duplicate or overlapping collector areas.')
            else:
                split_check_var.set(f'Collector expected is SHORT of the system total by {_fmt_amount(abs(diff))}. Add another collector or an Unassigned / Other row.')
            actual_var.set(f"{split_actual:.2f}")
        else:
            split_check_var.set('No collector split yet.')
        if select_index is not None and str(select_index) in split_tree.get_children():
            split_tree.selection_set(str(select_index))
            split_tree.focus(str(select_index))
        _recalc()
        _apply_mode()

    def _load_split_editor_from_selection(*_):
        idx = _selected_split_index()
        if idx is None:
            return
        rows = list(state.get('collector_rows') or [])
        if idx < 0 or idx >= len(rows):
            return
        row = rows[idx]
        state['selected_index'] = idx
        collector_name_var.set(str(row.get('collector_name') or ''))
        collector_expected_var.set(f"{float(row.get('expected_amount') or 0.0):.2f}")
        collector_actual_var.set(f"{float(row.get('actual_cash') or 0.0):.2f}")
        collector_note_var.set(str(row.get('note') or ''))

    def _upsert_split_row():
        if not can_edit_close:
            messagebox.showerror('Collector Split', 'Only the System user can edit collector cash split rows.')
            return
        rec = state.get('record') or {}
        if bool(int(rec.get('is_closed') or 0)):
            messagebox.showerror('Collector Split', 'Reopen the day first before changing collector split rows.')
            return
        name = (collector_name_var.get() or '').strip()
        if not name:
            messagebox.showwarning('Collector Split', 'Please enter a collector name.')
            return
        row = {
            'collector_name': name,
            'expected_amount': round(_parse_amount(collector_expected_var.get()), 2),
            'actual_cash': round(_parse_amount(collector_actual_var.get()), 2),
            'note': (collector_note_var.get() or '').strip(),
        }
        rows = list(state.get('collector_rows') or [])
        idx = state.get('selected_index')
        if idx is None:
            idx = _selected_split_index()
        if idx is None:
            rows.append(row)
            idx = len(rows) - 1
        else:
            rows[idx] = row
        for i, rr in enumerate(rows):
            rr['sort_order'] = i
        state['collector_rows'] = rows
        _refresh_split_tree(select_index=idx)
        _clear_split_editor()

    def _remove_split_row():
        if not can_edit_close:
            messagebox.showerror('Collector Split', 'Only the System user can edit collector cash split rows.')
            return
        rec = state.get('record') or {}
        if bool(int(rec.get('is_closed') or 0)):
            messagebox.showerror('Collector Split', 'Reopen the day first before changing collector split rows.')
            return
        idx = _selected_split_index()
        if idx is None:
            messagebox.showwarning('Collector Split', 'Please select a collector row to remove.')
            return
        rows = list(state.get('collector_rows') or [])
        if idx < 0 or idx >= len(rows):
            return
        rows.pop(idx)
        for i, rr in enumerate(rows):
            rr['sort_order'] = i
        state['collector_rows'] = rows
        _refresh_split_tree(select_index=max(0, idx - 1) if rows else None)
        _clear_split_editor()

    def _load_route_collectors():
        ds = (date_var.get() or '').strip()
        try:
            _dt.strptime(ds, '%Y-%m-%d')
        except Exception:
            messagebox.showerror('Collector Split', 'Use date format YYYY-MM-DD.')
            return
        rows = self._build_databank_collector_defaults_for_date(ds)
        if not rows:
            messagebox.showinfo('Collector Split', 'No collectors were found in collectors.json. You can still add rows manually.')
        state['collector_rows'] = rows
        _refresh_split_tree(select_index=0 if rows else None)
        _clear_split_editor()

    def _save_split():
        ds = (date_var.get() or '').strip()
        try:
            _dt.strptime(ds, '%Y-%m-%d')
        except Exception:
            messagebox.showerror('Collector Split', 'Use date format YYYY-MM-DD.')
            return
        if not can_edit_close:
            messagebox.showerror('Collector Split', 'Only the System user can save collector cash split rows.')
            return
        rows = list(state.get('collector_rows') or [])
        self.db.replace_databank_day_collectors(
            ds,
            rows,
            changed_by=(getattr(self, 'user_name', '') or '').strip(),
            source='databank:collector_split',
        )
        messagebox.showinfo('Collector Split', 'Collector split rows saved.')
        _load()

    def _apply_mode():
        rec = state.get('record') or {}
        is_closed = bool(int(rec.get('is_closed') or 0)) if rec else False
        has_split = bool(state.get('collector_rows'))
        lock_var.set('CLOSED' if is_closed else 'OPEN')
        edit_allowed = bool(can_edit_close)
        try:
            actual_ent.configure(state=('readonly' if (is_closed or (not edit_allowed) or has_split) else 'normal'))
        except Exception:
            pass
        try:
            note_ent.configure(state=('normal' if edit_allowed else 'disabled'))
        except Exception:
            pass
        try:
            workflow_cmb.configure(state=('readonly' if edit_allowed else 'disabled'))
        except Exception:
            pass
        try:
            close_btn.configure(state=('normal' if ((not is_closed) and edit_allowed) else 'disabled'))
        except Exception:
            pass
        try:
            reopen_btn.configure(state=('normal' if (is_closed and edit_allowed) else 'disabled'))
        except Exception:
            pass
        try:
            workflow_btn.configure(state=('normal' if edit_allowed else 'disabled'))
        except Exception:
            pass
        split_edit_state = ('normal' if ((not is_closed) and edit_allowed) else 'disabled')
        for b in (add_update_btn, remove_row_btn, load_routes_btn, save_split_btn):
            try:
                b.configure(state=split_edit_state)
            except Exception:
                pass
        try:
            use_split_btn.configure(state=('normal' if has_split else 'disabled'))
        except Exception:
            pass

    def _recalc(*_):
        expected = float(expected_box.get('value') or 0.0)
        if state.get('collector_rows'):
            actual = 0.0
            for row in (state.get('collector_rows') or []):
                try:
                    actual += float(row.get('actual_cash') or 0.0)
                except Exception:
                    pass
            actual = round(actual, 2)
            actual_var.set(f"{actual:.2f}")
        else:
            actual = _parse_amount(actual_var.get())
        variance = round(actual - expected, 2)
        if abs(variance) < 0.005:
            variance = 0.0
        variance_var.set(_fmt_amount(abs(variance)) if variance else _fmt_amount(0.0))
        status_var.set(_variance_status(variance))
        rec = state.get('record') or {}
        if not bool(int(rec.get('is_closed') or 0)):
            current_wf = (workflow_var.get() or '').strip()
            if current_wf not in ('Pending', 'Resolved'):
                workflow_var.set(_default_workflow(variance))

    def _load(*_):
        ds = (date_var.get() or '').strip()
        try:
            _dt.strptime(ds, '%Y-%m-%d')
        except Exception:
            messagebox.showerror('Daily Close', 'Use date format YYYY-MM-DD.')
            return
        mode_var.set('Combined (Regular + 7x7)')
        expected = round(float(self.db.get_databank_daily_total(ds, loan_type='__ALL__') or 0.0), 2)
        expected_box['value'] = expected
        expected_var.set(_fmt_amount(expected))
        rec = self.db.get_databank_day_close(ds)
        state['record'] = rec
        collector_rows = self.db.list_databank_day_collectors(ds)
        state['collector_rows'] = [
            {
                'collector_name': (r.get('collector_name') or '').strip(),
                'expected_amount': round(float(r.get('expected_amount') or 0.0), 2),
                'actual_cash': round(float(r.get('actual_cash') or 0.0), 2),
                'note': (r.get('note') or '').strip(),
                'sort_order': int(r.get('sort_order') or idx),
            }
            for idx, r in enumerate(collector_rows or []) if (r.get('collector_name') or '').strip()
        ]
        if rec:
            actual_var.set(f"{float(rec.get('actual_cash') or 0.0):.2f}")
            note_var.set((rec.get('note') or '').strip())
            try:
                variance = float(rec.get('variance') or 0.0)
            except Exception:
                variance = round(_parse_amount(actual_var.get()) - expected, 2)
            if abs(variance) < 0.005:
                variance = 0.0
            variance_var.set(_fmt_amount(abs(variance)) if variance else _fmt_amount(0.0))
            status_var.set((rec.get('variance_status') or _variance_status(variance)).strip() or 'Balanced')
            workflow_var.set((rec.get('variance_workflow_status') or _default_workflow(variance)).strip() or 'Open')
            who = (rec.get('closed_by') or '').strip()
            when = (rec.get('closed_at') or '').strip()
            open_who = (rec.get('opened_by') or '').strip()
            open_when = (rec.get('opened_at') or '').strip()
            if bool(int(rec.get('is_closed') or 0)):
                meta_var.set(f"Closed by {who or '—'} at {when or '—'}")
            else:
                meta_var.set(f"Open for editing. Last reopened by {open_who or '—'} at {open_when or '—'}")
        else:
            if state.get('collector_rows'):
                try:
                    actual_total = sum(float(r.get('actual_cash') or 0.0) for r in (state.get('collector_rows') or []))
                except Exception:
                    actual_total = expected
                actual_var.set(f"{actual_total:.2f}")
            else:
                actual_var.set(f"{expected:.2f}")
            note_var.set('')
            variance_var.set(_fmt_amount(0.0))
            status_var.set('Balanced')
            workflow_var.set('Resolved')
            meta_var.set('No combined close record yet for this date.')
        _refresh_split_tree(select_index=0 if state.get('collector_rows') else None)
        _apply_mode()
        try:
            self._update_data_toolbar()
        except Exception:
            pass
        _clear_split_editor()

    def _close_day():
        ds = (date_var.get() or '').strip()
        try:
            _dt.strptime(ds, '%Y-%m-%d')
        except Exception:
            messagebox.showerror('Daily Close', 'Use date format YYYY-MM-DD.')
            return
        if not can_edit_close:
            messagebox.showerror('Daily Close', 'Only the System user can close or change variance records.')
            return
        if not self._prompt_current_password(title='Close Day', prompt='Enter the System password to close this combined day and save the variance.'):
            return
        rows = list(state.get('collector_rows') or [])
        if rows:
            split_expected = sum(float(r.get('expected_amount') or 0.0) for r in rows)
            day_expected = float(expected_box.get('value') or 0.0)
            if abs(split_expected - day_expected) >= 0.005:
                messagebox.showerror('Daily Close', 'Collector expected total must match the system total before closing the day.')
                return
            actual = round(sum(float(r.get('actual_cash') or 0.0) for r in rows), 2)
        else:
            actual = _parse_amount(actual_var.get())
        variance = round(actual - float(expected_box.get('value') or 0.0), 2)
        if abs(variance) < 0.005:
            variance = 0.0
        workflow = (workflow_var.get() or '').strip() or _default_workflow(variance)
        rec = self.db.set_databank_day_close(
            ds,
            actual_cash=actual,
            note=(note_var.get() or '').strip(),
            closed_by=(getattr(self, 'user_name', '') or '').strip(),
            source='databank:close',
            workflow_status=workflow,
            collector_rows=(rows if rows else None),
        )
        state['record'] = rec
        _load()
        route_copy_path = ''
        try:
            if hasattr(self, 'save_closed_collector_route_copy'):
                route_copy_path = self.save_closed_collector_route_copy(
                    ds,
                    rec=rec,
                    collector_rows=(rows if rows else None),
                    open_after=False,
                ) or ''
        except Exception as e:
            route_copy_path = ''
            try:
                _log_exc('daily_close.save_collector_route_copy', e)
            except Exception:
                pass
        try:
            self.refresh_data_grid()
        except Exception:
            pass
        msg = f"{ds} is now closed for the combined Data Bank day (Regular + 7x7)."
        if route_copy_path:
            msg += f"\n\nCollector route close copy saved:\n{route_copy_path}"
        messagebox.showinfo('Daily Close', msg)

    def _reopen_day():
        ds = (date_var.get() or '').strip()
        try:
            _dt.strptime(ds, '%Y-%m-%d')
        except Exception:
            messagebox.showerror('Daily Close', 'Use date format YYYY-MM-DD.')
            return
        if not can_edit_close:
            messagebox.showerror('Daily Close', 'Only the System user can reopen or change variance records.')
            return
        rec = self.db.get_databank_day_close(ds)
        if not rec or not bool(int(rec.get('is_closed') or 0)):
            messagebox.showinfo('Daily Close', 'This day is already open.')
            _load()
            return
        if not self._prompt_current_password(title='Reopen Closed Day', prompt='Enter the System password to reopen this combined closed day for editing.'):
            return
        self.db.reopen_databank_day(
            ds,
            opened_by=(getattr(self, 'user_name', '') or '').strip(),
            source='databank:reopen',
        )
        _load()
        try:
            self.refresh_data_grid()
        except Exception:
            pass
        messagebox.showinfo('Daily Close', f"{ds} is open again. You can now edit Regular and 7x7 Data Bank entries for that day.")

    def _update_workflow():
        ds = (date_var.get() or '').strip()
        try:
            _dt.strptime(ds, '%Y-%m-%d')
        except Exception:
            messagebox.showerror('Daily Close', 'Use date format YYYY-MM-DD.')
            return
        if not can_edit_close:
            messagebox.showerror('Daily Close', 'Only the System user can change workflow or notes.')
            return
        rec = self.db.get_databank_day_close(ds)
        if not rec:
            messagebox.showerror('Daily Close', 'Load or close the day first before updating the workflow.')
            return
        if not self._prompt_current_password(title='Update Variance Workflow', prompt='Enter the System password to update workflow and note.'):
            return
        rec = self.db.set_databank_day_workflow(
            ds,
            workflow_var.get(),
            note=(note_var.get() or '').strip(),
            changed_by=(getattr(self, 'user_name', '') or '').strip(),
            source='databank:workflow',
        )
        state['record'] = rec
        _load()
        try:
            self.refresh_data_grid()
        except Exception:
            pass
        messagebox.showinfo('Daily Close', 'Variance workflow updated.')

    def _print_report():
        ds = (date_var.get() or '').strip()
        try:
            _dt.strptime(ds, '%Y-%m-%d')
        except Exception:
            messagebox.showerror('Daily Close', 'Use date format YYYY-MM-DD.')
            return
        self.print_databank_close_report(ds)

    def _save_route_copy_now():
        ds = (date_var.get() or '').strip()
        try:
            _dt.strptime(ds, '%Y-%m-%d')
        except Exception:
            messagebox.showerror('Daily Close', 'Use date format YYYY-MM-DD.')
            return
        try:
            rec = self.db.get_databank_day_close(ds) if hasattr(self, 'db') else None
        except Exception:
            rec = None
        if not rec or not bool(int((rec or {}).get('is_closed') or 0)):
            messagebox.showwarning('Save Route Copy', 'Close the day first before saving the final collector route copy.')
            return
        try:
            path = self.save_closed_collector_route_copy(ds, rec=rec, collector_rows=None, open_after=True)
        except Exception as e:
            try:
                _log_exc('daily_close.manual_save_route_copy', e)
            except Exception:
                pass
            messagebox.showerror('Save Route Copy', f'Could not save the collector route close copy.\n\n{e}')
            return
        messagebox.showinfo('Save Route Copy', f'Collector route close copy saved:\n{path}')

    def _open_records_list():
        self.open_databank_close_records_dialog(start_date=(date_var.get() or '').strip())

    load_btn.configure(command=_load)
    close_btn.configure(command=_close_day)
    reopen_btn.configure(command=_reopen_day)
    workflow_btn.configure(command=_update_workflow)
    print_btn.configure(command=_print_report)
    route_copy_btn.configure(command=_save_route_copy_now)
    list_btn.configure(command=_open_records_list)
    add_update_btn.configure(command=_upsert_split_row)
    remove_row_btn.configure(command=_remove_split_row)
    load_routes_btn.configure(command=_load_route_collectors)
    save_split_btn.configure(command=_save_split)
    use_split_btn.configure(command=lambda: actual_var.set(f"{sum(float(r.get('actual_cash') or 0.0) for r in (state.get('collector_rows') or [])):.2f}"))
    history_btn.configure(command=lambda: self.open_databank_close_history_dialog((date_var.get() or '').strip()))
    split_tree.bind('<<TreeviewSelect>>', _load_split_editor_from_selection)
    actual_var.trace_add('write', _recalc)
    _load()


def on_day_double(self, event):
    # Disabled in view-only mode
    return
    region = self.days_tree.identify('region', event.x, event.y)
    if region != 'cell': return
    row = self.days_tree.identify_row(event.y); col = self.days_tree.identify_column(event.x)
    if not row or not col: return
    col_idx = int(col.replace('#','')) - 1; day = col_idx + 1
    row_idx = self.days_tree.index(row)
    try: client_item = self.name_tree.get_children()[row_idx]; client = self.name_tree.item(client_item,'values')[0]
    except: return
    dt = date(self.grid_year, self.grid_month, day).strftime('%Y-%m-%d')
    bbox = self.days_tree.bbox(row, f'#{col_idx+1}');
    if not bbox: return
    x,y,w,h = bbox
    if self.current_entry:
        try: self.current_entry.destroy()
        except: pass
        self.current_entry = None
    ent = self._mk_tk_entry(self.days_tree, justify='center')
    cur = self.days_tree.set(row, f'd{day}').replace('','').replace(',','').strip()
    ent.insert(0, cur)
    ent.place(x=x, y=y, width=w, height=h); ent.focus_set(); self.current_entry = ent

    def save(e=None):
        v = ent.get().strip()
        # 1) Parse amount
        try:
            amt = float(v) if v != '' else 0.0
        except Exception:
            messagebox.showerror('Invalid', 'Enter a valid number.')
            return

        # 2) Persist to DB
        try:
            self.db.add_or_update_transaction(client, dt, amt, description='Daily Payment')
        except Exception as ex:
            messagebox.showerror('Save Error', f'Failed to save: {ex}')
            return

        # 3) Update UI cell (keep your existing zero-as-'0' display)
        try:
            self.days_tree.set(row, f'd{day}', fmt_currency(amt) if amt != 0 else '0')
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0395', 'suppressed exception excpass_0395', __spina_exc)
            pass

        # 4) Close editor
        try:
            ent.destroy()
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0396', 'suppressed exception excpass_0396', __spina_exc)
            pass
        self.current_entry = None

        # 5) Jump to next row for fast entry (wraps around)
        try:
            names = self.name_tree.get_children()
            if names:
                next_row = (row_idx + 1) % len(names)
                self.root.after(80, lambda: self._start_edit(next_row, col_idx))
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0397', 'suppressed exception excpass_0397', __spina_exc)
            pass

    # safer close helper to avoid double-destroy crashes
    def _close_editor(_=None):
        try:
            ent.destroy()
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0398', 'suppressed exception excpass_0398', __spina_exc)
            pass

    ent.bind('<Return>', save)
    ent.bind('<Escape>', _close_editor)
    ent.bind('<FocusOut>', _close_editor)


def _start_edit(self, row_idx, col_idx):
    try: item = self.days_tree.get_children()[row_idx]
    except: return
    day = col_idx+1
    bbox = self.days_tree.bbox(item, f'#{col_idx+1}')
    if not bbox:
        self.days_tree.see(item); self.name_tree.see(self.name_tree.get_children()[row_idx])
        self.days_tree.update_idletasks(); bbox = self.days_tree.bbox(item, f'#{col_idx+1}')
        if not bbox: return
    x,y,w,h = bbox
    client_item = self.name_tree.get_children()[row_idx]; client = self.name_tree.item(client_item,'values')[0]
    dt = date(self.grid_year, self.grid_month, day).strftime('%Y-%m-%d')
    if self.current_entry:
        try: self.current_entry.destroy()
        except: pass
        self.current_entry = None
    ent = self._mk_tk_entry(self.days_tree, justify='center')
    cur = self.days_tree.set(item, f'd{day}').replace('','').replace(',','').strip()
    ent.insert(0, cur)
    ent.place(x=x, y=y, width=w, height=h); ent.focus_set(); self.current_entry = ent

    def save(e=None):
        # 1) Parse amount
        try:
            txt = ent.get().strip()
            amt = float(txt) if txt != '' else 0.0
        except Exception:
            messagebox.showerror('Invalid', 'Enter a valid number.')
            return

        # 2) Save to DB
        try:
            self.db.add_or_update_transaction(client, dt, amt, description='Daily Payment')
        except Exception as ex:
            messagebox.showerror('Save Error', f'Failed to save: {ex}')
            return

        # 3) Update UI cell
        try:
            self.days_tree.set(item, f'd{day}', fmt_currency(amt) if amt != 0 else '0')
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0399', 'suppressed exception excpass_0399', __spina_exc)
            pass

        # 4) Close editor + advance to next row
        try:
            ent.destroy()
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0400', 'suppressed exception excpass_0400', __spina_exc)
            pass
        self.current_entry = None

        try:
            names = self.name_tree.get_children()
            if names:
                next_row = (row_idx + 1) % len(names)
                self.root.after(80, lambda: self._start_edit(next_row, col_idx))
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0401', 'suppressed exception excpass_0401', __spina_exc)
            pass

    def _close_editor(_=None):
        try:
            ent.destroy()
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0402', 'suppressed exception excpass_0402', __spina_exc)
            pass

    ent.bind('<Return>', save)
    ent.bind('<Escape>', _close_editor)
    ent.bind('<FocusOut>', _close_editor)


def _import_from_excel_entry(self):
    """Import payments from Excel.

        Supports:
          1) Date-grid templates (first row has 'Client Name' + date columns like YYYY-MM-DD) via _import_from_excel_core()
          2) One-day Daily Collection templates (Client | Payment | Reason), optionally grouped by [AREA] rows

        Notes:
          - Unknown clients are skipped (to avoid accidental new-client creation).
          - If a row has no payment but has a reason, we save a 0.0 payment with the reason.
        """
    from tkinter import filedialog, messagebox
    from datetime import datetime as _dt, date as _date
    import os, re

    # openpyxl required
    try:
        from openpyxl import load_workbook
    except Exception:
        try:
            messagebox.showerror("Missing dependency", "openpyxl is not installed. Install it to import Excel files.")
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0425', 'suppressed exception excpass_0425', __spina_exc)
            pass
        return

    paths = filedialog.askopenfilenames(
        title="Select import file(s)",
        filetypes=[
            ("Encoder batch / Excel", "*.jsonl;*.csv;*.xlsx;*.xlsm;*.xltx;*.xltm"),
            ("Encoder batch", "*.jsonl;*.csv"),
            ("Excel files", "*.xlsx;*.xlsm;*.xltx;*.xltm"),
            ("All files", "*.*"),
        ]
    )
    if not paths:
        return

    # Run import in a background thread to keep the UI responsive
    self._run_long_task('Importing...', lambda: self._import_from_excel_entry_worker(paths),
                        on_success=lambda _r: (self._ui_async(self.refresh_data_grid), self._ui_async(self.refresh_reports)))
    return


def _import_from_excel_entry_worker(self, paths):
    """Worker for _import_from_excel_entry (runs off the Tk main thread)."""
    from tkinter import messagebox
    from datetime import datetime as _dt, date as _date
    import os, re
    try:
        from openpyxl import load_workbook
    except Exception:
        try:
            messagebox.showerror('Missing dependency', 'openpyxl is not installed. Install it to import Excel files.')
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0426', 'suppressed exception excpass_0426', __spina_exc)
            pass
        return

    # If multiple files are selected, we import encoder batches one-by-one when applicable.
    # For Excel templates, we will import the first selected Excel file.
    path = paths[0]

    # Allow importing Encoder batch files (.jsonl / .csv)
    try:
        # If user selected multiple files, import all encoder batches
        imported_any = False
        for p in paths:
            try:
                ext = os.path.splitext(p)[1].lower()
            except Exception:
                ext = ""
            if ext in (".jsonl", ".csv"):
                self._import_encoder_batch(p)
                imported_any = True
        if imported_any:
            return
    except Exception:
        # fall back to single-file behavior below
        pass


    # loan type context
    try:
        lt = self._mode_filter()
    except Exception:
        lt = None

    # infer date from filename like DailyCollection_YYYY-MM-DD.xlsx
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", os.path.basename(path))
    inferred_date_str = f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m else _date.today().strftime("%Y-%m-%d")

    # Load workbook
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        try:
            messagebox.showerror("Open Error", f"Cannot open Excel file:\n{e}")
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0427', 'suppressed exception excpass_0427', __spina_exc)
            pass
        return

    def _norm(s):
        return (str(s or "").strip().lower()
                .replace("\u200b", "").replace("\xa0", " ").replace("  ", " "))

    def _looks_like_date_header(h):
        if h is None:
            return False
        if isinstance(h, (_dt, _date)):
            return True
        s = str(h).strip()
        # ISO date header
        try:
            _dt.fromisoformat(s)
            # Accept 'YYYY-MM-DD' or 'YYYY-MM-DD HH:MM:SS'
            return True
        except Exception:
            return False

    # Decide format using first sheet header row
    is_daily = False
    try:
        ws0 = wb.worksheets[0]
        hdr = next(ws0.iter_rows(min_row=1, max_row=1, values_only=True), None)
        labs = [_norm(x) for x in (hdr or [])]
        has_client = any(l in ("client", "client name", "name", "borrower", "customer") for l in labs)
        has_payment = any(l in ("payment", "amount", "paid", "amt") for l in labs)
        has_reason = any(l in ("reason", "remarks", "remark", "note", "notes", "comment", "comments") for l in labs)
        has_date_col = any(l in ("date", "collection date", "txn date") for l in labs)
        any_date_headers = any(_looks_like_date_header(x) for x in (hdr or []))

        # Daily collection template patterns:
        #   A) Client | Payment | Reason (no Date column, no date headers)
        #   B) Date | Client | Area | Payment | Reason
        if (has_client and has_payment and not any_date_headers) or (has_date_col and has_client and has_payment and not any_date_headers):
            is_daily = True

        # filename hint
        bn = os.path.basename(path).lower()
        if ("dailycollection" in bn or "daily_collection" in bn) and not any_date_headers:
            is_daily = True
    except Exception:
        # fallback: assume date-grid template
        is_daily = False

    # If not daily, delegate to the core importer (supports date-grid templates)
    if not is_daily:
        try:
            created, updated, skipped_rows, unknown = self._import_from_excel_core(path)
        except Exception as e:
            try:
                messagebox.showerror("Import Error", str(e))
            except Exception as __spina_exc:
                _log_suppressed_once('excpass_0428', 'suppressed exception excpass_0428', __spina_exc)
                pass
            try:
                wb.close()
            except Exception as __spina_exc:
                _log_suppressed_once('excpass_0429', 'suppressed exception excpass_0429', __spina_exc)
                pass
            return
        try:
            wb.close()
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0430', 'suppressed exception excpass_0430', __spina_exc)
            pass

        msg = f"Imported payments: {created} created, {updated} updated."
        details = []
        try:
            if unknown:
                details.append(f"Unknown clients: {len(unknown)}")
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0431', 'suppressed exception excpass_0431', __spina_exc)
            pass
        try:
            if skipped_rows:
                details.append(f"Skipped rows: {len(skipped_rows)}")
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0432', 'suppressed exception excpass_0432', __spina_exc)
            pass
        if details:
            msg += "\n" + " | ".join(details)
        try:
            messagebox.showinfo("Import Complete", msg)
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0433', 'suppressed exception excpass_0433', __spina_exc)
            pass
        try:
            self._ui_async(self.refresh_data_grid)
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0434', 'suppressed exception excpass_0434', __spina_exc)
            pass
        try:
            self._ui_async(self.refresh_reports)
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0435', 'suppressed exception excpass_0435', __spina_exc)
            pass
        return

    # ---- Daily collection import (one-day or explicit date column) ----
    HEADER_A = {
        "client": {"client", "client name", "name"},
        "payment": {"payment", "amount", "paid", "amt"},
        "reason": {"reason", "remarks", "remark", "note", "notes", "comment", "comments"},
    }
    HEADER_B = {
        "date": {"date", "collection date", "txn date"},
        "client": {"client", "client name", "name", "borrower", "customer"},
        "area": {"area", "zone", "barangay"},
        "payment": {"payment", "amount", "paid", "amt"},
        "reason": {"reason", "remarks", "remark", "note", "notes", "comment", "comments"},
    }

    records = []

    def _parse_sheet(ws):
        header_row = next((r for r in ws.iter_rows(min_row=1, max_row=1, values_only=True)), None)
        if not header_row:
            return

        colmap = {}
        labs = [_norm(x) for x in header_row]
        has_date = any(l in HEADER_B["date"] for l in labs)
        has_area_col = any(l in HEADER_B["area"] for l in labs)
        has_client = any(l in HEADER_A["client"] for l in labs)

        typeA = (has_client and not has_date and not has_area_col)

        if typeA:
            for idx, label in enumerate(header_row, start=1):
                lab = _norm(label)
                for canon, aliases in HEADER_A.items():
                    if lab in aliases:
                        colmap[idx] = canon
                        break
            if "client" not in colmap.values():
                return

            current_area = "UNASSIGNED"
            for row in ws.iter_rows(min_row=2, values_only=True):
                first = (str(row[0]).strip() if len(row) > 0 and row[0] is not None else "")
                m_area = re.match(r"^\[AREA\]\s*(.+)$", first, flags=re.IGNORECASE)
                if m_area:
                    current_area = (m_area.group(1) or "").strip().upper() or "UNASSIGNED"
                    continue
                if all((cell is None or str(cell).strip() == "") for cell in row):
                    continue

                def _get(canon):
                    for j, c in colmap.items():
                        if c == canon:
                            return row[j - 1]
                    return None

                name = str(_get("client") or "").strip()
                if not name:
                    continue

                raw_pay = _get("payment")
                raw_reason = _get("reason")
                pay = None
                if raw_pay is not None and str(raw_pay).strip() != "":
                    try:
                        pay = float(str(raw_pay).replace(",", "").strip())
                    except Exception:
                        pay = None
                reason = str(raw_reason).strip() if raw_reason is not None else ""

                # skip totally blank rows
                if (pay is None or pay == "") and not reason:
                    continue

                records.append({
                    "date": inferred_date_str,
                    "client": name,
                    "area": current_area,
                    "payment": pay,
                    "reason": reason,
                })

        else:
            for idx, label in enumerate(header_row, start=1):
                lab = _norm(label)
                for canon, aliases in HEADER_B.items():
                    if lab in aliases:
                        colmap[idx] = canon
                        break
            if not {"date", "client"}.issubset(set(colmap.values())):
                return

            for row in ws.iter_rows(min_row=2, values_only=True):
                if all((cell is None or str(cell).strip() == "") for cell in row):
                    continue

                def _get(canon):
                    for j, c in colmap.items():
                        if c == canon:
                            return row[j - 1]
                    return None

                raw_date = _get("date")
                raw_name = _get("client")
                raw_area = _get("area")
                raw_pay = _get("payment")
                raw_reason = _get("reason")

                name = str(raw_name or "").strip()
                if not name:
                    continue

                # Parse date
                dt = None
                if isinstance(raw_date, _dt):
                    dt = raw_date
                elif isinstance(raw_date, _date):
                    dt = _dt.combine(raw_date, _dt.min.time())
                else:
                    ds = str(raw_date or "").strip()
                    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%m-%d-%Y", "%d/%m/%Y", "%d-%m-%Y"):
                        try:
                            dt = _dt.strptime(ds, fmt)
                            break
                        except Exception as __spina_exc:
                            _log_suppressed_once('excpass_0436', 'suppressed exception excpass_0436', __spina_exc)
                            pass
                if not dt:
                    continue
                date_str = dt.strftime("%Y-%m-%d")

                area = str(raw_area).strip().upper() if raw_area is not None and str(raw_area).strip() != "" else ""

                pay = None
                if raw_pay is not None and str(raw_pay).strip() != "":
                    try:
                        pay = float(str(raw_pay).replace(",", "").strip())
                    except Exception:
                        pay = None
                reason = str(raw_reason).strip() if raw_reason is not None else ""

                if (pay is None or pay == "") and not reason:
                    continue

                records.append({
                    "date": date_str,
                    "client": name,
                    "area": area,
                    "payment": pay,
                    "reason": reason,
                })

    try:
        for ws in wb.worksheets:
            _parse_sheet(ws)
    finally:
        try:
            wb.close()
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0437', 'suppressed exception excpass_0437', __spina_exc)
            pass

    created = 0
    updated = 0
    unknown = set()

    for rec in records:
        name = rec.get("client", "").strip()
        ds = rec.get("date", "").strip()
        if not name or not ds:
            continue

        # unknown client check (mode-aware)
        try:
            info = self.db.get_client_info(name, loan_type=lt)
        except Exception:
            info = None
        if not info:
            unknown.add(name)
            continue

        # optionally fill missing area
        try:
            tmpl_area = (rec.get("area") or "").strip()
            existing_area = (info.get("area") or "").strip() if isinstance(info, dict) else ""
            if tmpl_area and not existing_area:
                try:
                    self.db.update_client(name, area=tmpl_area, loan_type=lt)
                except Exception as __spina_exc:
                    _log_suppressed_once('excpass_0438', 'suppressed exception excpass_0438', __spina_exc)
                    pass
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0439', 'suppressed exception excpass_0439', __spina_exc)
            pass

        pay = rec.get("payment", None)
        reason = (rec.get("reason") or "").strip()

        # If no payment but has a reason, store as 0.0
        if pay is None or (isinstance(pay, str) and pay.strip() == ""):
            if not reason:
                continue
            pay = 0.0

        # determine create vs update
        exists = False
        try:
            exists = self.db.get_transaction(name, ds, loan_type=lt) is not None
        except Exception:
            exists = False

        try:
            self.db.add_or_update_transaction(name, ds, float(pay or 0.0), description=reason, loan_type=lt)
        except Exception:
            continue

        if exists:
            updated += 1
        else:
            created += 1

    msg = f"Imported daily collection: {created} created, {updated} updated."
    if unknown:
        msg += f"\nUnknown clients skipped: {len(unknown)}"
    try:
        messagebox.showinfo("Import Complete", msg)
    except Exception as __spina_exc:
        _log_suppressed_once('excpass_0440', 'suppressed exception excpass_0440', __spina_exc)
        pass

    try:
        self._ui_async(self.refresh_data_grid)
    except Exception as __spina_exc:
        _log_suppressed_once('excpass_0441', 'suppressed exception excpass_0441', __spina_exc)
        pass
    try:
        self._ui_async(self.refresh_reports)
    except Exception as __spina_exc:
        _log_suppressed_once('excpass_0442', 'suppressed exception excpass_0442', __spina_exc)
        pass


def _import_encoder_batch(self, path: str):
    """Import One-Day Encoder exports (.jsonl or .csv) into the DB.

        - Dedupe by record_id (preferred) or a content hash fallback
        - Unknown clients are skipped (no auto-create)
        - Advances are stored via description tag:
            [ADV:s..e; s..e; ...]  (supports multiple ranges)
        """
    from tkinter import messagebox
    import os, json, csv, hashlib
    from datetime import datetime as _dt

    # Load import log (dedupe)
    os.makedirs(DATA_DIR, exist_ok=True)
    log_path = data_path( "encoder_import_log.json")
    try:
        with open(log_path, "r", encoding="utf-8") as f:
            log = json.load(f) or {}
    except Exception:
        log = {}

    def _canon_lt(v: str) -> str:
        s = (v or "").strip().lower().replace(" ", "")
        if "7x7" in s or "7×7" in s:
            return "7x7"
        return "Regular"

    def _is_date(s: str) -> bool:
        import re
        return bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}", s or ""))

    def _adv_tag_from_ranges(ranges):
        try:
            parts = []
            for it in (ranges or []):
                if isinstance(it, (list, tuple)) and len(it) >= 2:
                    a = str(it[0] or "").strip()
                    b = str(it[1] or "").strip()
                    if _is_date(a) and _is_date(b):
                        parts.append(f"{a}..{b}")
            if not parts:
                return ""
            return "[ADV:" + ";".join(parts) + "]"
        except Exception:
            return ""

    # Read records
    try:
        ext = os.path.splitext(path)[1].lower()
    except Exception:
        ext = ""
    recs = []
    try:
        if ext == ".jsonl":
            with open(path, "r", encoding="utf-8") as f:
                for line in f:
                    line = (line or "").strip()
                    if not line:
                        continue
                    try:
                        obj = json.loads(line)
                    except Exception:
                        continue
                    if isinstance(obj, dict):
                        recs.append(obj)
        elif ext == ".csv":
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if not isinstance(row, dict):
                        continue
                    recs.append(dict(row))
        else:
            messagebox.showerror("Import", "Unsupported file type. Please select .jsonl or .csv.")
            return
    except Exception as e:
        messagebox.showerror("Import failed", str(e))
        return

    if not recs:
        messagebox.showinfo("Import", "No rows found in the selected file.")
        return

    imported = 0
    skipped_dup = 0
    skipped_unknown = 0
    errors = 0
    inserted = 0
    updated = 0

    now_s = _dt.now().strftime("%Y-%m-%d %H:%M:%S")

    log_lines = []
    def _ilog(msg):
        try:
            log_lines.append(str(msg))
        except Exception:
            pass

    try:
        _ilog(f"FILE: {os.path.abspath(path)}")
    except Exception:
        _ilog(f"FILE: {path}")
    _ilog(f"START: {now_s} | Rows: {len(recs)}")
    _ilog("-" * 80)

    for i, r in enumerate(recs, start=1):
        try:
            # Dedupe/updates are handled after parsing (client/date/loan_type)

            raw_name = str(r.get("client") or r.get("name") or "").strip()
            if not raw_name:
                _ilog(f"[{i}] SKIP: missing client name")
                continue

            ds = str(r.get("encode_date") or r.get("date") or "").strip()
            if not _is_date(ds):
                _ilog(f"[{i}] SKIP: invalid date '{ds}' for raw_name='{raw_name}'")
                continue

            lt = _canon_lt(str(r.get("loan_type") or "Regular"))
            if hasattr(self.db, "_effective_lt"):
                try:
                    lt = self.db._effective_lt(lt)
                except Exception as __spina_exc:
                    _log_suppressed_once('excpass_0443', 'suppressed exception excpass_0443', __spina_exc)
                    pass

            # Prefer stable identifiers emitted by the Encoder export
            target_uid = str(r.get("target_client_uid") or r.get("client_uid") or "").strip()
            person_uid = str(r.get("person_uid") or "").strip()
            target_name_db = str(r.get("target_name_db") or "").strip()

            # Resolve the REAL DB client row by link (client_uid/person_uid) first, then fall back to name.
            resolved = None
            try:
                # 1) exact client_uid from export (best)
                if target_uid and hasattr(self.db, "get_client_by_uid"):
                    resolved = self.db.get_client_by_uid(target_uid)
                    # If export_uid points to the other loan_type, try by person_uid instead
                    try:
                        if resolved and (self.db._effective_lt(resolved.get("loan_type")) != lt):
                            resolved = None
                    except Exception:
                        pass

                # 2) person_uid link (Regular + 7x7 share one person_uid)
                if not resolved and person_uid:
                    try:
                        if hasattr(self.db, "get_client_by_person_uid_and_loan_type"):
                            resolved = self.db.get_client_by_person_uid_and_loan_type(person_uid, lt)
                    except Exception:
                        resolved = None
                    if not resolved and hasattr(self.db, "find_clients_by_person_uid"):
                        try:
                            for rr in (self.db.find_clients_by_person_uid(person_uid) or []):
                                try:
                                    if self.db._effective_lt(rr.get("loan_type")) == lt:
                                        resolved = rr
                                        break
                                except Exception:
                                    continue
                        except Exception:
                            resolved = None

                # 3) DB-targeted name emitted by export
                if not resolved and target_name_db:
                    try:
                        resolved = self.db.get_client_info(target_name_db, loan_type=lt)
                    except Exception:
                        resolved = None

                # 4) raw name (legacy fallback)
                if not resolved:
                    try:
                        resolved = self.db.get_client_info(raw_name, loan_type=lt)
                    except Exception:
                        resolved = None
            except Exception:
                resolved = None

            if not resolved:
                skipped_unknown += 1
                _ilog(f"[{i}] SKIP_UNKNOWN: raw_name='{raw_name}' date={ds} lt={lt} target_uid='{target_uid}' person_uid='{person_uid}' target_name_db='{target_name_db}'")
                continue

            # Use DB-canonical name/uid for all downstream operations
            name = str(resolved.get("name") or raw_name).strip()
            client_uid = str(resolved.get("client_uid") or target_uid).strip()

            # Stable import key per (date|loan_type|client_uid/name) so:
            # - Regular + 7x7 for the same person/day both import
            # - Re-importing an updated encoder export overwrites (upsert) correctly
            ek = str(r.get("export_key") or "").strip()
            if not ek:
                ek = f"{ds}|{lt}|{client_uid or name}"

            incoming_encoded_at_s = str(r.get("encoded_at") or r.get("encodedAt") or "").strip()

            def _parse_dt_safe(s: str):
                s = (s or "").strip()
                if not s:
                    return None
                try:
                    return _dt.fromisoformat(s)
                except Exception:
                    try:
                        return _dt.strptime(s, "%Y-%m-%d %H:%M:%S")
                    except Exception:
                        return None

            incoming_dt = _parse_dt_safe(incoming_encoded_at_s)
            prev = log.get(ek) if isinstance(log, dict) else None
            prev_dt = _parse_dt_safe(prev.get("encoded_at", "")) if isinstance(prev, dict) else None

            # Parse incoming payload before duplicate checks so we can compare against the DB.
            pay_raw = r.get("payment")
            try:
                pay = float(pay_raw) if str(pay_raw).strip() != "" else 0.0
            except Exception:
                pay = 0.0

            reason = str(r.get("reason") or r.get("description") or "").strip()

            # Encoder next-day route notice: stored separately from transactions so it does NOT
            # fill or affect tomorrow's encoder grid, but it can still print on the collector route PDF.
            route_notice = str(r.get("route_notice") or r.get("note_tomorrow") or r.get("next_route_notice") or "").strip()
            route_notice_date = str(r.get("route_notice_date") or r.get("next_notice_date") or "").strip()
            route_notice_source_date = str(r.get("route_notice_source_date") or ds or "").strip()
            if route_notice and route_notice_date:
                try:
                    if "_spina_route_notice_upsert" in globals():
                        _spina_route_notice_upsert(
                            day_iso=route_notice_date,
                            client=name,
                            loan_type=lt,
                            notice=route_notice,
                            client_uid=client_uid,
                            source_date=route_notice_source_date,
                            record_id=str(r.get("record_id") or ""),
                            collector=str(r.get("collector") or ""),
                        )
                except Exception as __spina_exc:
                    _log_suppressed_once('route_notice_import', 'route notice import failed', __spina_exc)

            adv_ranges = r.get("adv_ranges")
            # CSV may store lists as strings; try JSON decode
            if isinstance(adv_ranges, str) and adv_ranges.strip():
                try:
                    adv_ranges = json.loads(adv_ranges)
                except Exception:
                    adv_ranges = None

            adv_tag = _adv_tag_from_ranges(adv_ranges)

            desc = reason
            if adv_tag:
                desc = (desc + " " if desc else "") + adv_tag

            # Current DB row (prefer client_uid so renames don't create duplicates)
            current_tx = None
            try:
                if client_uid and hasattr(self.db, "get_transaction_by_uid"):
                    current_tx = self.db.get_transaction_by_uid(client_uid, ds, loan_type=lt)
                else:
                    current_tx = self.db.get_transaction(name, ds, loan_type=lt)
            except Exception:
                current_tx = None

            exists = current_tx is not None
            current_pay = None
            current_desc = ""
            try:
                if current_tx is not None:
                    if isinstance(current_tx, dict):
                        current_pay = float(current_tx.get("payment") or 0.0)
                        current_desc = str(current_tx.get("description") or "").strip()
                    elif hasattr(current_tx, "keys"):
                        current_pay = float(current_tx["payment"] or 0.0)
                        current_desc = str(current_tx["description"] or "").strip()
                    else:
                        current_pay = float((current_tx[4] if len(current_tx) > 4 else 0.0) or 0.0)
                        current_desc = str((current_tx[5] if len(current_tx) > 5 else "") or "").strip()
            except Exception:
                current_pay = None
                current_desc = ""

            same_as_db = False
            try:
                same_as_db = bool(exists) and (current_pay is not None) and (abs(float(current_pay) - float(pay)) < 0.00001) and (current_desc == desc)
            except Exception:
                same_as_db = False

            # If we have timestamps, skip only when the incoming record is NOT newer
            # AND the database still already contains the same payload. This allows
            # re-import after a row was deleted or blanked, even if the old import-log
            # entry still exists.
            if isinstance(prev, dict) and incoming_dt and prev_dt and incoming_dt <= prev_dt:
                if same_as_db:
                    skipped_dup += 1
                    try:
                        _ilog(f"[{i}] SKIP_DUP: ek='{ek}' incoming_encoded_at='{incoming_encoded_at_s}' prev_encoded_at='{(prev.get('encoded_at','') if isinstance(prev, dict) else '')}' uid='{client_uid}' name='{name}'")
                    except Exception:
                        pass
                    continue
                else:
                    try:
                        if isinstance(log, dict):
                            log.pop(ek, None)
                        _ilog(f"[{i}] STALE_DUP_LOG_CLEARED: ek='{ek}' uid='{client_uid}' name='{name}' date={ds} lt={lt}")
                    except Exception:
                        pass

            if client_uid and hasattr(self.db, "add_or_update_transaction_by_uid"):
                self.db.add_or_update_transaction_by_uid(client_uid, ds, pay, description=desc, loan_type=lt, source="encoder")
            else:
                self.db.add_or_update_transaction(name, ds, pay, description=desc, loan_type=lt, source="encoder")

            imported += 1
            inserted += 0 if exists else 1
            updated += 1 if exists else 0
            try:
                _ilog(f"[{i}] {'UPDATED' if exists else 'INSERTED'}: date={ds} lt={lt} uid='{client_uid}' name='{name}' payment={pay} desc='{desc}'")
            except Exception:
                pass
            log[ek] = {"imported_at": now_s, "encoded_at": (incoming_encoded_at_s or now_s), "record_id": str(r.get("record_id") or ""), "file": os.path.basename(path), "client": name, "date": ds, "loan_type": lt}
        except Exception as e:
            errors += 1
            try:
                _ilog(f"[{i}] ERROR: {repr(e)} raw_name='{(locals().get('raw_name','') or '')}' date='{(locals().get('ds','') or '')}' lt='{(locals().get('lt','') or '')}'")
            except Exception:
                pass
            continue
    try:
        if not _write_json_atomic(log_path, log):
            raise IOError('atomic write failed')
    except Exception as __spina_exc:
        _log_suppressed_once('import_log_write', 'import log write failed', __spina_exc)

    summary_msg = "\n".join([
        f"Imported: {imported}",
        f"Inserted: {inserted}",
        f"Updated: {updated}",
        f"Skipped duplicates: {skipped_dup}",
        f"Skipped unknown clients: {skipped_unknown}",
        f"Errors: {errors}",
    ])

    # Persist a human-readable log file per import (and keep a 'last' copy).
    log_txt_path = data_path("encoder_import_last.txt")
    try:
        ts = _dt.now().strftime("%Y%m%d_%H%M%S")
        log_txt_path = data_path(f"encoder_import_{ts}.txt")
        last_txt_path = data_path("encoder_import_last.txt")
        out_lines = []
        out_lines.append(summary_msg)
        out_lines.append("")
        out_lines.append("== CHRONOLOGICAL ==")
        out_lines.extend(log_lines or [])
        out_lines.append("")
        out_lines.append("== ORGANIZED ==")

        def _classify_line(_ln: str) -> str:
            s = (_ln or "").strip()
            up = s.upper()
            if "ERROR" in up and ("] ERROR" in up or "ERROR:" in up):
                return "Errors"
            if "SKIP_UNKNOWN" in up:
                return "Skipped Unknown"
            if "SKIP_DUP" in up:
                return "Skipped Duplicates"
            if "SKIP" in up and ("SKIP:" in up or "SKIP_" in up):
                return "Skipped"
            if "INSERTED" in up:
                return "Inserted"
            if "UPDATED" in up:
                return "Updated"
            if s.startswith("FILE:") or s.startswith("START:") or s.startswith("END:") or s.startswith("-" * 10):
                return "Header/Info"
            return "Other"

        _cats = ["Inserted", "Updated", "Skipped Duplicates", "Skipped Unknown", "Skipped", "Errors", "Header/Info", "Other"]
        _grp = {c: [] for c in _cats}
        for _ln in (log_lines or []):
            try:
                c = _classify_line(str(_ln))
                _grp.setdefault(c, []).append(str(_ln))
            except Exception:
                _grp.setdefault("Other", []).append(str(_ln))

        for c in _cats:
            sec = _grp.get(c) or []
            if not sec:
                continue
            out_lines.append(f"-- {c} ({len(sec)}) --")
            out_lines.extend(sec)
            out_lines.append("")

        for pth in (log_txt_path, last_txt_path):
            try:
                with open(pth, "w", encoding="utf-8") as f:
                    for ln in out_lines:
                        f.write(str(ln) + "\n")
            except Exception as __spina_exc:
                _log_suppressed_once('excpass_importlog_write', 'suppressed exception excpass_importlog_write', __spina_exc)
    except Exception as __spina_exc:
        _log_suppressed_once('excpass_importlog_write2', 'suppressed exception excpass_importlog_write2', __spina_exc)

    try:
        messagebox.showinfo("Import complete", summary_msg)
    except Exception as __spina_exc:
        _log_suppressed_once('excpass_0444', 'suppressed exception excpass_0444', __spina_exc)
        pass

    # Show detailed log viewer
    try:
        self._show_import_log_window("Encoder Import Log", summary_msg + "\nSaved: " + str(log_txt_path), log_lines, default_save_path=log_txt_path)
    except Exception as __spina_exc:
        _log_suppressed_once('excpass_importlog_ui', 'suppressed exception excpass_importlog_ui', __spina_exc)
        pass


def _import_from_excel_core(self, path, progress_cb=None):
    """
        Auto-detect payments and reasons from Excel.

        Rules:
          " First row = headers. Must include 'Client Name' and at least one date column.
          " A date column can be an Excel date/datetime or a string 'YYYY-MM-DD'.
          " If a cell under a date column is numeric (or numeric-looking text): that's the amount.
          " If a cell under a date column is non-numeric text: that's a REASON (description).
          " If there are sibling columns whose header mentions the same date plus
            'reason/remark/remarks/note' (any case), we also append their text as reason.
          " We always save the reason (description) when present. If no amount is present,
            we write a zero-amount transaction with the description so it shows in Data Bank.
        "
          " If neither amount nor reason is present, we now CLEAR the existing value by writing 0.0 with empty reason.
        """
    from openpyxl import load_workbook
    import datetime as _dt

    def _parse_date_header(h):
        if h is None:
            return None
        # Excel might give datetime/date objects already
        if isinstance(h, (_dt.datetime, _dt.date)):
            return h.date() if isinstance(h, _dt.datetime) else h
        s = str(h).strip()
        # expect headers like 'YYYY-MM-DD'
        try:
            return _dt.date.fromisoformat(s)
        except Exception:
            return None

    def _to_number(val):
        if val is None:
            return None
        # allow numeric-like strings with commas
        if isinstance(val, str):
            s = val.strip()
            if not s:
                return None
            # if it looks like a number after removing commas, parse it
            try:
                return float(s.replace(",", ""))
            except Exception:
                return None
        # genuine numbers
        if isinstance(val, (int, float)):
            return float(val)
        return None

    def _is_text_reason(val):
        if val is None:
            return False
        # numeric-looking? then it's NOT a reason
        if _to_number(val) is not None:
            return False
        # any non-empty non-numeric content counts as a reason
        return str(val).strip() != ""

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [ (c.value) for c in ws[1] ]
    if not headers or not any(h is not None and str(h).strip() for h in headers):
        wb.close()
        raise ValueError("The first row must contain headers (Client Name and at least one date).")

    # Locate 'Client Name' (fallback to first column if not found)
    low = [str(h).strip().lower() if h is not None else "" for h in headers]
    c_name = low.index("client name") if "client name" in low else 0

    # Build date columns and discover possible sibling reason columns per date
    # Example headers: "2025-10-01", "2025-10-01 Reason", "2025-10-01 remarks"
    date_cols = []                # list[(col_index, date_obj, date_str)]
    date_reason_map = {}          # col_index -> [reason_col_indices]
    header_strs = [str(h).strip() if h is not None else "" for h in headers]
    for idx, h in enumerate(headers):
        if idx == c_name:
            continue
        d = _parse_date_header(h)
        if d:
            date_cols.append((idx, d, d.isoformat()))
            date_reason_map[idx] = []

    # For each date header, find any "companion" reason columns that mention the same date and contain reason-ish keywords.
    reason_keywords = ("reason", "remark", "remarks", "note", "notes")
    for amt_idx, _, d_str in date_cols:
        d_lower = d_str.lower()
        for j, hj in enumerate(header_strs):
            if j == amt_idx:
                continue
            hj_low = hj.lower()
            if d_lower in hj_low and any(k in hj_low for k in reason_keywords):
                date_reason_map[amt_idx].append(j)

    # Known clients map (case-insensitive)
    existing = {r[0].strip().lower(): r[0] for r in self.db.conn.execute("SELECT name FROM clients")}

    created, updated = 0, 0
    skipped_rows = []
    unknown = []

    # Count rows for optional progress
    row_iter = list(ws.iter_rows(min_row=2, values_only=True))
    total_rows = len(row_iter)
    for r_i, row in enumerate(row_iter, start=1):
        if progress_cb:
            try:
                progress_cb(r_i, total_rows)
            except Exception as __spina_exc:
                _log_suppressed_once('excpass_0445', 'suppressed exception excpass_0445', __spina_exc)
                pass

        if not row:
            continue
        raw_name = row[c_name] if c_name < len(row) else None
        if raw_name is None or str(raw_name).strip() == "":
            continue

        key = str(raw_name).strip().lower()
        if key not in existing:
            unknown.append(str(raw_name).strip())
            continue
        name = existing[key]

        for amt_idx, d, d_str in date_cols:
            if amt_idx >= len(row):
                continue

            amount = _to_number(row[amt_idx])
            reason_parts = []

            # If the "amount cell" actually contains text (non-numeric), treat it as reason text
            if _is_text_reason(row[amt_idx]):
                reason_parts.append(str(row[amt_idx]).strip())

            # Pull any companion reason/remarks cells for this same date
            for r_idx in date_reason_map.get(amt_idx, []):
                if r_idx < len(row) and _is_text_reason(row[r_idx]):
                    reason_parts.append(str(row[r_idx]).strip())

            # Compose reason text (dedupe + join)
            reason_txt = "; ".join(dict.fromkeys([p for p in reason_parts if p])) if reason_parts else ""

            # If neither amount nor reason is present, CLEAR (write 0.0 with empty reason)
            if amount is None and not reason_txt:
                try:
                    self.db.add_or_update_transaction(name, d.isoformat(), 0.0, description="")
                    updated += 1
                except Exception as e:
                    skipped_rows.append((name, d.isoformat(), str(e)))
                continue

            # If amount is None but we have reason, write a 0.0 entry with description
            final_amt = amount if amount is not None else 0.0

            try:
                # Always keep the reason text (even when amount > 0)
                self.db.add_or_update_transaction(name, d.isoformat(), final_amt, description=reason_txt or "Imported")
                # We don't know if your add_or_update returns info; we count all as created for simplicity
                created += 1
            except Exception as e:
                skipped_rows.append((name, d.isoformat(), str(e)))

    wb.close()
    return created, updated, skipped_rows, unknown


def import_from_excel_with_reasons(self):
    """
    Import the Excel 'range template' where columns are:
    Client Name | 2025-10-01 | 2025-10-01 Reason | 2025-10-02 | 2025-10-02 Reason | ...
    Saves both Amount and Reason for each date.
    """
    from tkinter import filedialog, messagebox
    try:
        from openpyxl import load_workbook
    except Exception:
        try:
            messagebox.showwarning("Missing", "Install openpyxl to import from Excel.")
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0583', 'suppressed exception excpass_0583', __spina_exc)
            pass
        return

    path = filedialog.askopenfilename(
        title="Select Excel to Import",
        filetypes=[("Excel Workbook", "*.xlsx")]
    )
    if not path:
        return

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        try:
            messagebox.showerror("Open Error", f"Cannot open file:\\n{e}")
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0584', 'suppressed exception excpass_0584', __spina_exc)
            pass
        return

    ws = wb.active
    # --- Read header row
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        headers.append(str(v).strip() if v is not None else "")

    # Validate first header
    if not headers or (headers[0] or "").strip().lower() != "client name":
        try:
            messagebox.showerror("Format Error", "First header must be 'Client Name'.")
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0585', 'suppressed exception excpass_0585', __spina_exc)
            pass
        return

    # Build map: date_col_index -> (date_str, reason_col_index or None)
    header_to_index = {h: i+1 for i, h in enumerate(headers) if h}
    date_cols = {}
    from datetime import datetime as _dt
    for idx, h in enumerate(headers, start=1):
        if idx == 1 or not h:
            continue
        ht = h.strip()
        # detect YYYY-MM-DD
        try:
            _dt.strptime(ht, "%Y-%m-%d")
            is_date = True
        except Exception:
            is_date = False
        if is_date:
            reason_idx = header_to_index.get(f"{ht} Reason")
            date_cols[idx] = (ht, reason_idx)

    if not date_cols:
        try:
            messagebox.showerror("Format Error", "No date columns found (e.g., 'YYYY-MM-DD').")
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0586', 'suppressed exception excpass_0586', __spina_exc)
            pass
        return

    inserted = 0
    skipped  = 0
    errors   = 0

    # Choose a DB write function
    def _write(name, ymd, amount, reason):
        nonlocal inserted, skipped, errors
        try:
            # Parse amount if numeric, otherwise treat text as a reason
            amt = None
            reason_txt = (str(reason).strip() if reason is not None else "")

            if amount not in (None, ""):
                s = str(amount).strip()
                try:
                    amt = float(s)              # numeric amount
                except Exception:
                    # Text typed in the Amount cell -> treat as reason, set amount to 0
                    if s:
                        reason_txt = (f"{reason_txt}; {s}".strip("; ").strip()) if reason_txt else s
                        amt = 0.0

            # If nothing to save, skip
            if (amt is None or amt == 0.0) and not reason_txt:
                skipped += 1
                return

            # Normalize "Advance" ranges into [ADV:YYYY-MM-DD..YYYY-MM-DD]
            try:
                import re as _re
                m = _re.search(r"(?:ADV|Advance)[^0-9]*"
                               r"(\d{4}-\d{2}-\d{2}).*?(\d{4}-\d{2}-\d{2})",
                               reason_txt, flags=_re.I)
                if m:
                    s_norm, e_norm = m.group(1), m.group(2)
                    tag = f"[ADV:{s_norm}..{e_norm}]"
                    if tag not in reason_txt:
                        reason_txt = f"{reason_txt} {tag}".strip()
            except Exception as __spina_exc:
                _log_suppressed_once('excpass_0587', 'suppressed exception excpass_0587', __spina_exc)
                pass

            # Write to DB: keep description only when payment is 0/blank
            if hasattr(self.db, "add_or_update_transaction"):
                self.db.add_or_update_transaction(
                    name, ymd,
                    (amt if amt is not None else 0.0),
                    description=reason_txt
                )
            elif hasattr(self.db, "add_payment"):
                self.db.add_payment(name, ymd, (amt if amt is not None else 0.0), reason=reason_txt)
            elif hasattr(self.db, "insert_payment"):
                self.db.insert_payment(name, ymd, (amt if amt is not None else 0.0), reason=reason_txt)
            else:
                conn = getattr(self.db, "conn", None)
                if not conn:
                    raise RuntimeError("No DB connection to insert payment.")
                cur = conn.cursor()
                try:
                    cur.execute(
                        "INSERT INTO payments (name, date, amount, reason) VALUES (?,?,?,?)",
                        (name, ymd, (amt if amt is not None else 0.0), (reason_txt or None))
                    )
                except Exception:
                    cur.execute(
                        "INSERT INTO transactions (name, date, payment, description) VALUES (?,?,?,?)",
                        (name, ymd, (amt if amt is not None else 0.0), (reason_txt or None))
                    )
                conn.commit()

            inserted += 1
        except Exception:
            errors += 1


def _spina_perf_refresh_data_grid(self):
    """Fast Data Bank month grid refresh using bulk month transaction query."""
    try:
        self.status_var.set("Data Bank is view-only. Encode in Excel, then Import.")
    except Exception:
        pass

    try:
        for w in self.inner.winfo_children():
            try:
                w.destroy()
            except Exception:
                pass
    except Exception:
        return

    days = calendar.monthrange(self.grid_year, self.grid_month)[1]
    day_cols = tuple(f"d{d}" for d in range(1, days + 1))
    cols = ("client", "area") + day_cols

    grid = ttk.Frame(self.inner)
    grid.grid(row=0, column=0, sticky="nsew")
    try:
        self.inner.grid_rowconfigure(0, weight=1)
        self.inner.grid_columnconfigure(0, weight=1)
    except Exception:
        pass
    grid.grid_rowconfigure(0, weight=1)
    grid.grid_columnconfigure(1, weight=1)

    self.name_tree = ttk.Treeview(grid, columns=("client", "area"), show="headings", height=31)
    self._configure_tree_stripes(self.name_tree)
    self.name_tree.heading("client", text="Client Name")
    self.name_tree.heading("area", text="Area")
    self.name_tree.column("client", width=330, anchor="w", stretch=False)
    self.name_tree.column("area", width=185, anchor="w", stretch=False)

    self.days_tree = ttk.Treeview(grid, columns=cols, show="headings", height=31)
    self._configure_tree_stripes(self.days_tree)
    self.days_tree.heading("client", text="")
    self.days_tree.heading("area", text="")
    self.days_tree.column("client", width=0, minwidth=0, stretch=False)
    self.days_tree.column("area", width=0, minwidth=0, stretch=False)
    for c in day_cols:
        d = int(c[1:])
        self.days_tree.heading(c, text=str(d))
        self.days_tree.column(c, width=86, anchor="center", stretch=False)

    def _yview(*args):
        try: self.name_tree.yview(*args)
        except Exception: pass
        try: self.days_tree.yview(*args)
        except Exception: pass

    v = ttk.Scrollbar(grid, orient="vertical", command=_yview)
    h = ttk.Scrollbar(grid, orient="horizontal", command=self.days_tree.xview)
    self.name_tree.configure(yscrollcommand=v.set)
    self.days_tree.configure(yscrollcommand=v.set, xscrollcommand=h.set)
    self.name_tree.grid(row=0, column=0, sticky="ns")
    self.days_tree.grid(row=0, column=1, sticky="nsew")
    v.grid(row=0, column=2, sticky="ns")
    h.grid(row=1, column=1, sticky="ew")

    def _sync_selection(from_tv, to_tv):
        try:
            sel = from_tv.selection()
            if not sel:
                return
            if to_tv.selection() != sel:
                to_tv.selection_set(sel)
            try:
                to_tv.focus(sel[0])
            except Exception:
                pass
        except Exception:
            pass

    try:
        self.name_tree.bind("<<TreeviewSelect>>", lambda e=None: _sync_selection(self.name_tree, self.days_tree), add="+")
        self.days_tree.bind("<<TreeviewSelect>>", lambda e=None: _sync_selection(self.days_tree, self.name_tree), add="+")
        self.days_tree.bind("<Double-1>", self._begin_cell_edit, add="+")
        self.days_tree.bind("<Button-1>", self._remember_cell_click, add="+")
        self.days_tree.bind("<Delete>", self.delete_selected_cell, add="+")
        self.name_tree.bind("<Delete>", self.delete_selected_cell, add="+")
        for tv in (self.name_tree, self.days_tree):
            tv.bind("<MouseWheel>", self._on_mousewheel_sync, add="+")
            tv.bind("<Button-4>", self._on_mousewheel_sync, add="+")
            tv.bind("<Button-5>", self._on_mousewheel_sync, add="+")
    except Exception as e:
        try:
            _log_ignored("ui.bind failed", e, key="ui.bind_failed")
        except Exception:
            pass

    try:
        self._db_menu = tk.Menu(self.days_tree, tearoff=0)
        self._db_menu.add_command(label="Mark as missed (enter reason)", command=self._mark_missed_for_selected)
        self._db_menu.add_separator()
        self._db_menu.add_command(label="Delete this payment cell", command=self.delete_selected_cell)
        def _popup_db_menu(ev):
            try: self._remember_cell_click(ev)
            except Exception: pass
            try:
                self._db_menu.tk_popup(ev.x_root, ev.y_root)
            finally:
                try: self._db_menu.grab_release()
                except Exception: pass
        self.days_tree.bind("<Button-3>", _popup_db_menu, add="+")
    except Exception:
        pass

    try:
        self.inner.bind("<Configure>", self._resize_databank_columns, add="+")
    except Exception:
        pass

    try:
        search_term = self.search_db_var.get().strip()
    except Exception:
        search_term = ""
    try:
        mode_txt = self._mode_filter()
    except Exception:
        mode_txt = "Regular"

    # Data Bank should show current loan type only. No 7x7 extras here.
    try:
        clients = _spina_perf_clients_rows(self.db, loan_type=mode_txt, search=search_term if search_term else None, search_by="all", include_extra_7x7=False)
    except Exception as e:
        try:
            _log_exc("perf_refresh_data_grid: clients bulk load failed", e)
        except Exception:
            pass
        clients = []

    try:
        rows_txt = f"{len(clients or [])} row{'s' if len(clients or []) != 1 else ''} • {mode_txt} • {self._month_label()}"
        if search_term:
            rows_txt += f" • filter: {search_term}"
        if hasattr(self, "_db_rows_var") and self._db_rows_var is not None:
            self._db_rows_var.set(rows_txt)
    except Exception:
        pass

    if not clients:
        iid = "r0"
        self.name_tree.insert("", "end", iid=iid, values=("(no clients)", ""), tags=("odd",))
        self.days_tree.insert("", "end", iid=iid, values=("(no clients)", "", *([''] * days)), tags=("odd",))
    else:
        start_date = date(self.grid_year, self.grid_month, 1).strftime("%Y-%m-%d")
        end_date = date(self.grid_year, self.grid_month, days).strftime("%Y-%m-%d")
        try:
            payment_map = _spina_perf_month_transactions(self.db, clients, start_date, end_date, mode_txt)
        except Exception as e:
            try:
                _log_exc("perf_refresh_data_grid: tx bulk load failed", e)
            except Exception:
                pass
            payment_map = {}

        for idx, info in enumerate(clients):
            try:
                name = str(info.get("name") or "")
                area = str(info.get("area") or "")
                uid = str(info.get("client_uid") or "").strip()
                key = uid if uid else name.strip().lower()
                vals = [name, area]
                for d in range(1, days + 1):
                    ds = date(self.grid_year, self.grid_month, d).strftime("%Y-%m-%d")
                    if (key, ds) in payment_map:
                        p = payment_map.get((key, ds))
                        try:
                            if p is not None and float(p) != 0:
                                vals.append(fmt_currency(p))
                            else:
                                vals.append("0")
                        except Exception:
                            vals.append(str(p or ""))
                    else:
                        vals.append("")
                tag = "odd" if (idx % 2) == 0 else "even"
                iid = f"r{idx}"
                try:
                    self.name_tree.insert("", "end", iid=iid, values=(name, area), tags=(tag,))
                except Exception:
                    self.name_tree.insert("", "end", values=(name, area), tags=(tag,))
                try:
                    self.days_tree.insert("", "end", iid=iid, values=tuple(vals), tags=(tag,))
                except Exception:
                    self.days_tree.insert("", "end", values=tuple(vals), tags=(tag,))
            except Exception as e:
                try:
                    _log_suppressed_once("perf_databank_row_insert", "databank row insert skipped", e)
                except Exception:
                    pass

    try:
        self._resize_databank_columns()
    except Exception:
        pass
    try:
        if hasattr(self, "_update_data_toolbar"):
            self._update_data_toolbar()
    except Exception:
        pass


def _spina_auto_close_one_day(db, ds, days_after):
    """Close one Data Bank day using system expected totals as the safe automatic actual amount."""
    expected = 0.0
    try:
        expected = round(float(db.get_databank_daily_total(ds, loan_type='__ALL__') or 0.0), 2)
    except Exception:
        expected = 0.0
    if expected <= 0:
        return None

    actual = expected
    collector_rows_for_save = None
    try:
        collector_rows = db.list_databank_day_collectors(ds)
    except Exception:
        collector_rows = []

    # If a collector split exists and its expected total matches the day total,
    # preserve it and use the collector actual cash total. Otherwise close as balanced.
    try:
        if collector_rows:
            split_expected = round(sum(float(r.get('expected_amount') or 0.0) for r in collector_rows), 2)
            if abs(split_expected - expected) < 0.005:
                actual = round(sum(float(r.get('actual_cash') or 0.0) for r in collector_rows), 2)
                collector_rows_for_save = collector_rows
    except Exception:
        actual = expected
        collector_rows_for_save = None

    variance = round(actual - expected, 2)
    if abs(variance) < 0.005:
        variance = 0.0
    workflow = 'Resolved' if variance == 0 else 'Pending'
    note = f"Auto close after {int(days_after)} day(s)."
    return db.set_databank_day_close(
        ds,
        actual_cash=actual,
        note=note,
        closed_by='Auto Close',
        source='databank:auto_close',
        workflow_status=workflow,
        collector_rows=collector_rows_for_save,
    )


def _spina_run_auto_daily_close(self, show_message=False):
    """Auto-close Data Bank dates that are older than the configured delay."""
    from datetime import date as _date, timedelta as _timedelta
    try:
        days_after = _spina_auto_close_after_days_value()
    except Exception:
        days_after = 0
    if days_after <= 0:
        if show_message:
            try:
                messagebox.showinfo('Auto Daily Close', 'Auto Daily Close is disabled. Set days to 1 or higher in Settings.')
            except Exception:
                pass
        return {'closed': 0, 'skipped': 0, 'days_after': days_after}

    if not hasattr(self, 'db') or self.db is None:
        return {'closed': 0, 'skipped': 0, 'days_after': days_after}

    cutoff = _date.today() - _timedelta(days=int(days_after))
    cutoff_s = cutoff.isoformat()
    closed = []
    skipped = []
    try:
        candidates = _spina_auto_close_candidate_dates(self.db, cutoff_s)
        for ds in candidates:
            try:
                rec = _spina_auto_close_one_day(self.db, ds, days_after)
                if rec:
                    closed.append(ds)
                    try:
                        if hasattr(self, 'save_closed_collector_route_copy'):
                            self.save_closed_collector_route_copy(ds, rec=rec, collector_rows=None, open_after=False)
                    except Exception as e:
                        try:
                            _log_exc(f'auto_daily_close.save_collector_route_copy.{ds}', e)
                        except Exception:
                            pass
                else:
                    skipped.append(ds)
            except Exception as e:
                skipped.append(ds)
                try:
                    _log_exc(f'auto_daily_close.day.{ds}', e)
                except Exception:
                    pass
    except Exception as e:
        try:
            _log_exc('auto_daily_close.run', e)
        except Exception:
            pass
        if show_message:
            try:
                messagebox.showerror('Auto Daily Close', str(e))
            except Exception:
                pass
        return {'closed': len(closed), 'skipped': len(skipped), 'days_after': days_after}

    try:
        if closed and hasattr(self, 'refresh_data_grid'):
            self.refresh_data_grid()
    except Exception:
        pass
    try:
        if closed and hasattr(self, '_update_data_toolbar'):
            self._update_data_toolbar()
    except Exception:
        pass

    if show_message:
        try:
            if closed:
                sample = ', '.join(closed[:8])
                more = f" and {len(closed) - 8} more" if len(closed) > 8 else ''
                messagebox.showinfo('Auto Daily Close', f"Closed {len(closed)} day(s):\n{sample}{more}")
            else:
                messagebox.showinfo('Auto Daily Close', f"No due days to auto-close.\nCurrent setting: after {days_after} day(s).")
        except Exception:
            pass
    return {'closed': len(closed), 'skipped': len(skipped), 'days_after': days_after}


def _spina_save_closed_collector_route_copy(self, date_s, rec=None, collector_rows=None, open_after=False):
    """Save an audit copy of the Collector Route after Daily Close.

    The PDF contains the closed total/actual cash for the day and the amount paid by each
    client on that date. It is separate from Generate Report and from editable Data Bank rows.
    """
    import os as _os
    from datetime import datetime as _dt
    try:
        from reportlab.pdfgen import canvas as _canvas
        from reportlab.lib.pagesizes import A4, landscape as _landscape
        from reportlab.lib import colors as _colors
    except Exception as e:
        raise RuntimeError("ReportLab is required to save the closed collector route PDF.") from e

    ds = str(date_s or "").strip()[:10]
    try:
        _dt.strptime(ds, "%Y-%m-%d")
    except Exception:
        raise ValueError("Use date format YYYY-MM-DD.")

    if rec is None:
        try:
            rec = self.db.get_databank_day_close(ds) if hasattr(self, "db") else None
        except Exception:
            rec = None
    if not rec:
        raise ValueError("No Daily Close record exists yet for that date.")

    def _rec_get(k, default=None):
        try:
            return rec.get(k, default)
        except Exception:
            try:
                return dict(rec).get(k, default)
            except Exception:
                return default

    try:
        expected = round(float(_rec_get("expected_amount", 0.0) or 0.0), 2)
    except Exception:
        expected = 0.0
    try:
        actual = round(float(_rec_get("actual_cash", 0.0) or 0.0), 2)
    except Exception:
        actual = 0.0
    try:
        variance = round(float(_rec_get("variance", actual - expected) or 0.0), 2)
    except Exception:
        variance = round(actual - expected, 2)
    if abs(variance) < 0.005:
        variance = 0.0
    try:
        reg_total = round(float(getattr(self, "db").get_databank_daily_total(ds, loan_type="Regular") or 0.0), 2)
    except Exception:
        reg_total = 0.0
    try:
        x7_total = round(float(getattr(self, "db").get_databank_daily_total(ds, loan_type="7x7") or 0.0), 2)
    except Exception:
        x7_total = 0.0

    collectors = _spina_crc_load_collectors()
    rows = _spina_crc_fetch_close_rows(self, ds, collectors)

    # Use saved collector split rows from the close record when available.
    close_collectors = []
    if collector_rows is not None:
        try:
            close_collectors = [dict(r) for r in (collector_rows or [])]
        except Exception:
            close_collectors = []
    if not close_collectors:
        try:
            close_collectors = list(self.db.list_databank_day_collectors(ds) or [])
        except Exception:
            close_collectors = []
    close_by_name = {}
    for r in close_collectors or []:
        try:
            nm = str((r or {}).get("collector_name") or "").strip()
            if nm:
                close_by_name[nm.lower()] = dict(r)
        except Exception:
            pass

    base_dir = ""
    try:
        base_dir = data_path("Closed_Collector_Routes", ds)
    except Exception:
        base_dir = _os.path.join(DATA_DIR if 'DATA_DIR' in globals() else _os.getcwd(), "Closed_Collector_Routes", ds)
    _os.makedirs(base_dir, exist_ok=True)
    stamp = _dt.now().strftime("%H%M%S")
    paid_tag = _safe_filename_component(_spina_crc_fmt_money(actual).replace(",", ""), fallback="0") if '_safe_filename_component' in globals() else str(int(actual or 0))
    out_path = _os.path.join(base_dir, f"ClosedCollectorRoute_{ds}_Paid_{paid_tag}_{stamp}.pdf")

    # Also copy any previously printed collector-route PDFs for the same date into the close folder.
    copied_route_pdfs = _spina_crc_copy_existing_route_pdfs(ds, base_dir)

    page_w, page_h = _landscape(A4)
    c = _canvas.Canvas(out_path, pagesize=(page_w, page_h))
    margin = 24
    y = page_h - margin
    line_h = 10.5

    col_no = 18
    col_area = 82
    col_name = 180
    col_type = 38
    col_paid = 70
    col_reason = page_w - (margin * 2) - (col_no + col_area + col_name + col_type + col_paid)

    def _new_page():
        nonlocal y
        c.showPage()
        y = page_h - margin
        _draw_header(small=True)

    def _ensure(h):
        if y - h < margin + 22:
            _new_page()

    def _draw_header(small=False):
        nonlocal y
        c.setFillColor(_colors.black)
        c.setFont("Helvetica-Bold", 13 if not small else 11)
        c.drawString(margin, y, f"CLOSED COLLECTOR ROUTE COPY - {ds}")
        c.setFont("Helvetica", 8.5)
        c.drawRightString(page_w - margin, y, "Saved after Daily Close")
        y -= 14
        if not small:
            c.setFont("Helvetica", 9)
            summary = (
                f"Amount Paid Today: { _spina_crc_fmt_money(actual) }    "
                f"Expected: { _spina_crc_fmt_money(expected) }    "
                f"Variance: { _spina_crc_fmt_money(variance) }    "
                f"Regular: { _spina_crc_fmt_money(reg_total) }    7x7: { _spina_crc_fmt_money(x7_total) }"
            )
            c.drawString(margin, y, summary)
            y -= 12
            closed_by = str(_rec_get("closed_by", "") or "").strip()
            closed_at = str(_rec_get("closed_at", "") or "").strip()
            wf = str(_rec_get("variance_workflow_status", "") or "").strip()
            c.drawString(margin, y, f"Closed by: {closed_by or '-'}    Closed at: {closed_at or '-'}    Workflow: {wf or '-'}")
            y -= 14
            if copied_route_pdfs:
                c.setFillColor(_colors.HexColor("#555555"))
                c.drawString(margin, y, f"Existing printed route PDF copy/copies saved in the same folder: {len(copied_route_pdfs)}")
                c.setFillColor(_colors.black)
                y -= 12
        c.setStrokeColor(_colors.HexColor("#999999"))
        c.line(margin, y, page_w - margin, y)
        y -= 10

    def _draw_table_header():
        nonlocal y
        _ensure(18)
        c.saveState()
        c.setFillColor(_colors.HexColor("#F2F2F2"))
        c.rect(margin, y - 13, page_w - (margin * 2), 15, stroke=0, fill=1)
        c.restoreState()
        c.setFont("Helvetica-Bold", 8)
        x = margin + 2
        for title, w in [("#", col_no), ("Area", col_area), ("Client", col_name), ("Type", col_type), ("Paid", col_paid), ("Reason / detail", col_reason)]:
            if title == "Paid":
                c.drawRightString(x + w - 3, y - 8, title)
            else:
                c.drawString(x, y - 8, title)
            x += w
        y -= 17

    _draw_header(small=False)

    if not rows:
        c.setFont("Helvetica", 10)
        c.drawString(margin, y, "No route/payment rows found for this date.")
        y -= 16
    else:
        # Keep collector order from collectors.json, then append any extra/unassigned groups.
        order = [str(c0.get("name") or "").strip() for c0 in collectors if str(c0.get("name") or "").strip()]
        groups = {}
        for r in rows:
            groups.setdefault(str(r.get("collector") or "Unassigned"), []).append(r)
        for g in list(groups.keys()):
            if g not in order:
                order.append(g)

        for collector in order:
            gr = groups.get(collector) or []
            if not gr and collector.lower() not in close_by_name:
                continue
            sum_paid = round(sum(float(r.get("paid") or 0.0) for r in gr), 2)
            close_rec = close_by_name.get(str(collector).lower(), {})
            try:
                close_actual = round(float((close_rec or {}).get("actual_cash") or 0.0), 2)
            except Exception:
                close_actual = 0.0
            try:
                close_expected = round(float((close_rec or {}).get("expected_amount") or 0.0), 2)
            except Exception:
                close_expected = 0.0
            _ensure(34)
            c.setFillColor(_colors.HexColor("#EAF2F8"))
            c.rect(margin, y - 16, page_w - (margin * 2), 18, stroke=0, fill=1)
            c.setFillColor(_colors.black)
            c.setFont("Helvetica-Bold", 9.5)
            title = f"Collector: {collector}"
            c.drawString(margin + 4, y - 10, title)
            c.setFont("Helvetica", 8.5)
            if close_rec:
                right = f"Closed actual: {_spina_crc_fmt_money(close_actual)}   Expected: {_spina_crc_fmt_money(close_expected)}   Route rows paid: {_spina_crc_fmt_money(sum_paid)}"
            else:
                right = f"Route rows paid: {_spina_crc_fmt_money(sum_paid)}"
            c.drawRightString(page_w - margin - 4, y - 10, right)
            y -= 24
            _draw_table_header()
            idx = 1
            for r in gr:
                area = str(r.get("area") or "-")
                nm = str(r.get("name") or "")
                lt = _spina_crc_norm_lt(r.get("loan_type") or "Regular")
                paid = _spina_crc_fmt_money(r.get("paid") or 0.0, blank_zero=True)
                reason = str(r.get("reason") or r.get("detail") or "")
                if r.get("detail") and reason:
                    if str(r.get("detail")) not in reason:
                        reason = (reason + " | " + str(r.get("detail"))).strip(" |")
                elif r.get("detail"):
                    reason = str(r.get("detail") or "")
                name_lines = _spina_crc_wrap(c, nm, col_name - 4, "Helvetica", 8.2)
                reason_lines = _spina_crc_wrap(c, reason, col_reason - 4, "Helvetica", 7.4)
                row_lines = max(1, len(name_lines), len(reason_lines))
                h = max(14, 8 + (row_lines * line_h))
                _ensure(h + 2)
                if idx % 2 == 0:
                    c.saveState()
                    c.setFillColor(_colors.HexColor("#FBFBFB"))
                    c.rect(margin, y - h + 2, page_w - (margin * 2), h, stroke=0, fill=1)
                    c.restoreState()
                c.setStrokeColor(_colors.HexColor("#DDDDDD"))
                c.line(margin, y - h + 2, page_w - margin, y - h + 2)
                c.setFillColor(_colors.black)
                c.setFont("Helvetica", 8)
                x = margin + 2
                c.drawString(x, y - 8, str(idx)); x += col_no
                area_lines = _spina_crc_wrap(c, area, col_area - 4, "Helvetica", 7.6)
                for i, ln in enumerate(area_lines[:row_lines]):
                    c.drawString(x, y - 8 - (i * line_h), ln)
                x += col_area
                c.setFont("Helvetica-Bold", 8.2)
                for i, ln in enumerate(name_lines):
                    c.drawString(x, y - 8 - (i * line_h), ln)
                x += col_name
                c.setFont("Helvetica", 8)
                c.drawString(x, y - 8, "7x7" if lt == "7x7" else "REG")
                x += col_type
                c.setFont("Helvetica-Bold", 8)
                c.drawRightString(x + col_paid - 4, y - 8, paid)
                x += col_paid
                c.setFont("Helvetica", 7.4)
                for i, ln in enumerate(reason_lines[:max(1, row_lines)]):
                    c.drawString(x, y - 8 - (i * line_h), ln)
                y -= h
                idx += 1
            y -= 6

    _ensure(24)
    c.setFont("Helvetica-Oblique", 7.5)
    c.setFillColor(_colors.HexColor("#666666"))
    c.drawString(margin, margin - 2 + 10, "This PDF was saved automatically when the Data Bank day was closed. Amounts are from that close date's payment records.")
    c.drawRightString(page_w - margin, margin - 2 + 10, f"Saved: {_dt.now().strftime('%Y-%m-%d %H:%M:%S')}")
    c.save()

    try:
        self._last_closed_collector_route_copy_path = out_path
    except Exception:
        pass
    if open_after:
        try:
            _open_path(out_path)
        except Exception:
            pass
    return out_path
