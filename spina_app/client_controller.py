from __future__ import annotations

import json
import os
import re
import shutil
import sqlite3
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Mapping

import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

_CLIENT_CONTROLLER_DEPENDENCIES: dict[str, Any] = {}


def configure_client_controller_dependencies(namespace: Mapping[str, Any]) -> None:
    _CLIENT_CONTROLLER_DEPENDENCIES.clear()
    _CLIENT_CONTROLLER_DEPENDENCIES.update(namespace)
    protected = {"__name__", "__file__", "__package__", "__builtins__", "_CLIENT_CONTROLLER_DEPENDENCIES", "configure_client_controller_dependencies"}
    for name, value in namespace.items():
        if name not in protected:
            globals()[name] = value


def set_area_for_selected_clients(self):
        """Set one area for all selected clients in the Clients tab (current mode only)."""
        try:
            sel = list(self.clients_tree.selection())
        except Exception:
            sel = []
        if not sel:
            messagebox.showwarning('Select', 'Please select one or more clients first.')
            return

        area = (getattr(self, 'bulk_area_var', tk.StringVar(value='')).get() or '').strip()
        try:
            areas = set(self.db.get_all_areas() or [])
        except Exception:
            areas = set()

        if area and areas and (area not in areas):
            messagebox.showerror('Invalid', 'Selected area is not in your Areas list. Please add it first.')
            return

        updated = 0
        for iid in sel:
            try:
                name = self.clients_tree.item(iid, 'values')[0]
            except Exception:
                continue
            if not name:
                continue

            # Infer actual loan_type from the row (supports 7x7-only rows shown in Regular view)
            lt = self._mode_filter()
            try:
                tags = self.clients_tree.item(iid, 'tags') or ()
                for t in tags:
                    if str(t).startswith('lt:'):
                        lt = str(t).split(':', 1)[1]
                        break
            except Exception as __spina_exc:
                _log_suppressed_once('excpass_0458', 'suppressed exception excpass_0458', __spina_exc)
                pass

            try:
                self.db.update_client(name, area=area, loan_type=lt)
                updated += 1
            except Exception as __spina_exc:
                _log_suppressed_once('excpass_0459', 'suppressed exception excpass_0459', __spina_exc)
                pass

        self.refresh_clients()
        self.refresh_reports()
        self.refresh_data_grid()
        messagebox.showinfo('Done', f'Updated area for {updated} client(s).')

def _app__get_selected_client_name(self):
    try:
        sel = self.clients_tree.selection()
        if not sel:
            return None
        vals = self.clients_tree.item(sel[0], 'values') or ()
        if not vals:
            return None
        return str(vals[0]).strip()
    except Exception:
        return None

def _app_refresh_clients(self):
    try:
        term = (getattr(self, "search_clients_var", tk.StringVar()).get() or "").strip()
    except Exception:
        term = ""
    lt = _app__norm_lt_value(self, self._mode_filter())
    try:
        for rid in self.clients_tree.get_children():
            self.clients_tree.delete(rid)
    except Exception:
        return

    try:
        try:
            mode = (getattr(self, "clients_search_mode_var", tk.StringVar(value="All")).get() or "Both").strip().lower()
        except Exception:
            mode = "both"
        if mode.startswith("cli"):
            sb = "client"
        elif mode.startswith("are"):
            sb = "area"
        elif mode.startswith("link"):
            sb = "linked"
        elif mode.startswith("unl"):
            sb = "unlinked"
        elif mode.startswith("sug"):
            sb = "suggested"
        elif mode.startswith("blank"):
            sb = "blanks"
        elif mode.startswith("pri"):
            sb = "principal"
        elif mode.startswith("rel"):
            sb = "released"
        elif mode.startswith("sta"):
            sb = "start_date"
        elif mode.startswith("due"):
            sb = "due_date"
        else:
            sb = "all"
        names = self.db.get_all_clients(search=term if term else None, loan_type=lt, search_by=sb) or []
    except Exception:
        names = []

    other_lt = _app__other_lt(self, lt)

    # Build rows to show:
    # - Always show the current mode's clients
    # - If viewing Regular: append 7x7-only (no Regular record and not linked) at the bottom
    rows_to_show = []
    for n in (names or []):
        rows_to_show.append((n, lt, False))

    if str(lt) == "Regular" and sb not in ("linked", "suggested"):
        try:
            extra_7x7 = self.db.get_unpaired_7x7_names(search=term if term else None, search_by=sb) or []
        except Exception:
            extra_7x7 = []
        for n in (extra_7x7 or []):
            # Avoid duplicates (safety)
            if any(str(n).strip().lower() == str(a[0]).strip().lower() for a in rows_to_show):
                continue
            rows_to_show.append((n, "7x7", True))

    for name, row_lt, is_extra_7x7 in rows_to_show:
        info = self.db.get_client_info(name, loan_type=row_lt) or {}
        link_label = ""
        try:
            if is_extra_7x7:
                link_label = "7x7 ONLY"
            else:
                row_other_lt = _app__other_lt(self, row_lt)
                pu = (info.get("person_uid") or "").strip()
                if pu:
                    linked = self.db.find_clients_by_person_uid(pu) or []
                    has_other = any(str(r.get("loan_type")) != row_lt for r in linked)
                    link_label = f"{row_other_lt}✓" if has_other else "Linked✓"
                else:
                    if self.db.get_client_info(name, loan_type=row_other_lt):
                        link_label = row_other_lt
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0631', 'suppressed exception excpass_0631', __spina_exc)
            pass

        # Store the real loan_type on the row so edits/deletes/area changes work even when 7x7-only is shown in Regular view
        _day_due, _due_today = _spina__client_due_meta(info)
        _day_due_val = _day_due
        row_no = len(self.clients_tree.get_children())
        tags = [f"lt:{row_lt}", 'even' if (row_no % 2 == 0) else 'odd']
        if is_extra_7x7:
            tags.append('extra7x7')
        iid = None
        try:
            iid = (info.get("client_uid") or "").strip() or None
        except Exception:
            iid = None

        row_values = (
            info.get("name", name),
            info.get("area", ""),
            info.get("payment_term", ""),
            _day_due_val,
            _spina__fmt_client_money(info.get("payment_amount", 0)),
            info.get("payment_mode", "Cash"),
            link_label,
            info.get("contact_number", ""),
            _spina__fmt_client_money(info.get("principal", 0)),
            _spina__fmt_client_money(info.get("interest_amount", 0)),
            _spina__fmt_client_money(info.get("total_to_pay", 0)),
            info.get("date_released", ""),
            info.get("due_date", ""),
        )

        try:
            self.clients_tree.insert(
                "", "end",
                iid=iid,
                tags=tuple(tags),
                values=row_values,
            )
        except Exception:
            # Fallback insert if iid duplicates or fails
            try:
                self.clients_tree.insert(
                    "", "end",
                    tags=tuple(tags),
                    values=row_values,
                )
            except Exception as __spina_exc:
                _log_suppressed_once('excpass_0632', 'suppressed exception excpass_0632', __spina_exc)
                pass

    try:
        if hasattr(self, 'clients_count_var'):
            self.clients_count_var.set(f"Rows: {len(self.clients_tree.get_children())}")
    except Exception as __spina_exc:
        _log_suppressed_once('excpass_0632b', 'suppressed exception excpass_0632b', __spina_exc)
        pass

    try:
        self._update_toolbar_states()
    except Exception as __spina_exc:
        _log_suppressed_once('excpass_0633', 'suppressed exception excpass_0633', __spina_exc)
        pass
    # Refresh area dropdowns only when NOT actively filtering (prevents lag/jitter while typing)
    if not term:
        try:
            self._refresh_area_dropdowns()
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0634', 'suppressed exception excpass_0634', __spina_exc)
            pass

def _app_schedule_refresh_clients(self, *_):
    """Debounce Clients search so refresh doesn't run on every keystroke."""
    root = getattr(self, 'root', None) or getattr(self, 'master', None)
    # Cancel previous scheduled refresh, if any
    try:
        after_id = getattr(self, '_clients_refresh_after_id', None)
        if after_id and root is not None:
            root.after_cancel(after_id)
    except Exception as __spina_exc:
        _log_suppressed_once('excpass_0635', 'suppressed exception excpass_0635', __spina_exc)
        pass
    # Schedule a new refresh shortly
    try:
        if root is not None:
            self._clients_refresh_after_id = root.after(150, self.refresh_clients)
        else:
            self.refresh_clients()
    except Exception:
        try:
            self.refresh_clients()
        except Exception as __spina_exc:
            _log_suppressed_once('excpass_0636', 'suppressed exception excpass_0636', __spina_exc)
            pass

            pass

def _app_delete_client_selected(self):
    name = _app__get_selected_client_name(self)
    if not name:
        return
    lt = _app__norm_lt_value(self, self._mode_filter())
    if not messagebox.askyesno("Confirm", f"Archive (hide) client '{name}' ({lt})?\n\n• It will disappear\n• History stays\n• You can restore anytime"):
        return
    try:
        ok = self.db.archive_client(name, loan_type=lt)
        if not ok:
            messagebox.showerror("Failed", "Delete failed.")
            return
    except Exception as e:
        messagebox.showerror("Error", str(e))
        return
    try:
        self.refresh_clients()
        self.refresh_reports()
        self.refresh_data_grid()
    except Exception as __spina_exc:
        _log_suppressed_once('excpass_0641', 'suppressed exception excpass_0641', __spina_exc)
        pass
    messagebox.showinfo("Archived", f"Client '{name}' archived (hidden).")

def _app_link_selected_client(self):
    name = _app__get_selected_client_name(self)
    if not name:
        messagebox.showinfo("Link", "Select a client first.")
        return
    lt = _app__norm_lt_value(self, self._mode_filter())
    other_lt = _app__other_lt(self, lt)

    uid_a = self.db.get_client_uid(name, loan_type=lt)
    if not uid_a:
        messagebox.showerror("Link", "Could not get selected client UID.")
        return

    win = tk.Toplevel(self.root)
    win.title(f"Link to {other_lt} client")
    win.geometry("520x480")
    try:
        win.transient(self.root)
        win.grab_set()
    except Exception as __spina_exc:
        _log_suppressed_once('excpass_0644', 'suppressed exception excpass_0644', __spina_exc)
        pass

    top = ttk.Frame(win, padding=8)
    top.pack(fill='both', expand=True)

    ttk.Label(top, text=f"Choose {other_lt} record to link with:").pack(anchor='w')
    qv = tk.StringVar()
    ttk.Entry(top, textvariable=qv).pack(fill='x', pady=6)

    lb = tk.Listbox(top)
    lb.pack(fill='both', expand=True)

    # Selected row itself must be unlinked before linking.
    try:
        _cur_info0 = self.db.get_client_info(name, loan_type=lt) or {}
        if (_cur_info0.get('person_uid') or '').strip():
            messagebox.showinfo("Already Linked", f"'{name}' ({lt}) is already linked. Unlink it first.")
            try:
                win.destroy()
            except Exception:
                pass
            return
    except Exception:
        pass

    all_names = []
    try:
        for _nm in (self.db.get_all_clients(loan_type=other_lt) or []):
            _info = self.db.get_client_info(_nm, loan_type=other_lt) or {}
            if not (_info.get('person_uid') or '').strip():
                all_names.append(_nm)
    except Exception:
        all_names = []

    def refresh_list():
        q = (qv.get() or "").strip().lower()
        lb.delete(0, 'end')
        for nm in all_names:
            if not q or q in nm.lower():
                lb.insert('end', nm)

    qv.trace_add('write', lambda *a: refresh_list())
    refresh_list()

    if not all_names:
        messagebox.showinfo("No records", f"No unlinked {other_lt} clients exist yet.", parent=win)

    def do_link():
        try:
            idx = lb.curselection()
            if not idx:
                messagebox.showinfo("Link", "Select a record to link.", parent=win)
                return
            target_name = lb.get(idx[0])
            uid_b = self.db.get_client_uid(target_name, loan_type=other_lt)
            if not uid_b:
                messagebox.showerror("Link", "Could not get target UID.", parent=win)
                return
            cur_info = self.db.get_client_info(name, loan_type=lt) or {}
            other_info = self.db.get_client_info(target_name, loan_type=other_lt) or {}
            pu_a = (cur_info.get('person_uid') or '').strip()
            pu_b = (other_info.get('person_uid') or '').strip()
            if pu_a and pu_b and pu_a == pu_b:
                messagebox.showinfo("Already Linked", f"'{name}' ({lt}) and '{target_name}' ({other_lt}) are already linked.", parent=win)
                return
            if pu_a or pu_b:
                who = name if pu_a else target_name
                messagebox.showinfo("Already Linked", f"'{who}' is already linked. Unlink it first.", parent=win)
                return
            ok = self.db.link_client_uids(uid_a, uid_b, source='clients:manual_link')
            if not ok:
                err = ''
                try:
                    err = self.db.get_last_error()
                except Exception:
                    err = ''
                if err and 'already linked' in err.lower():
                    messagebox.showinfo("Already Linked", err, parent=win)
                else:
                    messagebox.showerror("Link", err or "Link failed.", parent=win)
                return
            try:
                win.destroy()
            except Exception as __spina_exc:
                _log_suppressed_once('excpass_0645', 'suppressed exception excpass_0645', __spina_exc)
                pass
            self.refresh_clients()
            messagebox.showinfo("Linked", f"Linked '{name}' ({lt}) with '{target_name}' ({other_lt}).")
        except Exception as e:
            messagebox.showerror("Link", str(e), parent=win)

    btn = ttk.Frame(top)
    btn.pack(fill='x', pady=8)
    ttk.Button(btn, text="Cancel", command=lambda: win.destroy()).pack(side='right')
    ttk.Button(btn, text="Link", command=do_link).pack(side='right', padx=8)

    try:
        win.bind('<Return>', lambda e: do_link())
        win.bind('<Escape>', lambda e: win.destroy())
    except Exception as e:
        _log_ignored("ui.bind failed", e, key="ui.bind_failed")

def _app_unlink_selected_client(self):
    name = _app__get_selected_client_name(self)
    if not name:
        messagebox.showinfo("Unlink", "Select a client first.")
        return
    lt = _app__norm_lt_value(self, self._mode_filter())
    info = self.db.get_client_info(name, loan_type=lt) or {}
    pu = (info.get("person_uid") or "").strip()
    if not pu:
        messagebox.showinfo("Unlink", "Selected client is not linked.")
        return
    if not messagebox.askyesno("Confirm", "Unlink this client group (Regular + 7x7)?"):
        return
    try:
        ok = self.db.unlink_person_uid(pu, source='clients:manual_unlink')
        if ok:
            self.refresh_clients()
            messagebox.showinfo("Unlinked", "Unlinked successfully.")
        else:
            messagebox.showerror("Failed", "Unlink failed.")
    except Exception as e:
        messagebox.showerror("Error", str(e))

def _app__maybe_suggest_link_clients(self, name, loan_type):
    try:
        nm = (name or "").strip()
        if not nm:
            return
        lt = _app__norm_lt_value(self, loan_type or self._mode_filter())
        other_lt = _app__other_lt(self, lt)
        other_info = self.db.get_client_info(nm, loan_type=other_lt)
        if not other_info:
            return
        meta_a = self.db.get_client_link_meta(nm, loan_type=lt) or {}
        meta_b = self.db.get_client_link_meta(nm, loan_type=other_lt) or {}
        uid_a = meta_a.get("client_uid")
        uid_b = meta_b.get("client_uid")
        if not uid_a or not uid_b:
            return
        if int(meta_a.get("link_opt_out") or 0) == 1 or int(meta_b.get("link_opt_out") or 0) == 1:
            return
        pu_a = (meta_a.get("person_uid") or "").strip()
        pu_b = (meta_b.get("person_uid") or "").strip()
        if pu_a and pu_b and pu_a == pu_b:
            return
        msg = (
            f"'{nm}' exists in BOTH loan types:\n\n"
            f"• {lt}\n"
            f"• {other_lt}\n\n"
            "Link them as the SAME person?\n\n"
            "YES = connect (still separate loans)\n"
            "NO = keep separate and stop asking again for this pair"
        )
        if messagebox.askyesno("Link suggestion", msg):
            ok = False
            try:
                ok = self.db.link_client_uids(uid_a, uid_b, source='clients:auto_suggest:yes')
            except Exception:
                ok = False
            if not ok:
                err = ''
                try:
                    err = self.db.get_last_error()
                except Exception:
                    err = ''
                messagebox.showerror("Link", err or "Link failed.")
                return
            try:
                self.refresh_clients()
            except Exception as __spina_exc:
                _log_suppressed_once('excpass_0646', 'suppressed exception excpass_0646', __spina_exc)
                pass
        else:
            try:
                self.db.set_link_opt_out_by_uid(uid_a, 1, source='clients:auto_suggest:no')
            except Exception as __spina_exc:
                _log_suppressed_once('excpass_0647', 'suppressed exception excpass_0647', __spina_exc)
                pass
            try:
                self.db.set_link_opt_out_by_uid(uid_b, 1, source='clients:auto_suggest:no')
            except Exception as __spina_exc:
                _log_suppressed_once('excpass_0648', 'suppressed exception excpass_0648', __spina_exc)
                pass
    except Exception:
        return

def _app_export_clients_template(self):
    try:
        import openpyxl
        from openpyxl import Workbook
    except Exception:
        messagebox.showerror("Excel", "Missing dependency: openpyxl. Install with: pip install openpyxl")
        return
    path = filedialog.asksaveasfilename(
        title="Save Clients Template",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile="clients_template.xlsx",
    )
    if not path:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"
    headers = ["Name", "Contact Number", "Loan Type (Regular/7x7)", "Area", "Principal", "Interest Rate (Regular only)", "Date Released (YYYY-MM-DD)", "New Until (optional)", "Pay Start Offset Days (0/1)"]
    ws.append(headers)
    try:
        wb.save(path)
        messagebox.showinfo("Saved", f"Template saved:\n{path}")
    except Exception as e:
        messagebox.showerror("Save Error", str(e))

def _app_import_clients_from_excel(self):
    try:
        import openpyxl
    except Exception:
        messagebox.showerror("Excel", "Missing dependency: openpyxl. Install with: pip install openpyxl")
        return
    path = filedialog.askopenfilename(title="Select Clients Excel", filetypes=[("Excel", "*.xlsx")])
    if not path:
        return
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    except Exception as e:
        messagebox.showerror("Excel", str(e))
        return

    lt_default = _app__norm_lt_value(self, self._mode_filter())
    imported = 0
    updated = 0
    failed = 0
    failure_samples = []

    rows_iter = iter(ws.iter_rows(values_only=True))
    first_row = next(rows_iter, None)
    if first_row is None:
        messagebox.showinfo("Import", "No data found.")
        return
    header = [str(x).strip().lower() if x is not None else "" for x in first_row]

    def col_idx(keys):
        for k in keys:
            if k in header:
                return header.index(k)
        return None

    i_name = col_idx(["name"])
    i_contact = col_idx(["contact number", "contact no", "contact", "phone", "mobile", "cell", "cp number", "tel", "telephone"])
    i_lt = col_idx(["loan type (regular/7x7)", "loan type", "loantype"])
    i_area = col_idx(["area"])
    i_pr = col_idx(["principal"])
    i_ir = col_idx(["interest rate (regular only)", "interest rate", "interestrate"])
    i_dr = col_idx(["date released (yyyy-mm-dd)", "date released", "released", "date_released"])
    i_nu = col_idx(["new until (optional)", "new until", "new_until"])
    i_ps = col_idx(["pay start offset days (0/1)", "pay start offset", "pay_start_offset_days", "offset"])

    if i_name is None:
        messagebox.showerror("Import", "Template must include a 'Name' column.")
        return

    for row_number, r in enumerate(rows_iter, start=2):
        try:
            nm = (r[i_name] or "")
            nm = " ".join(str(nm).strip().split())
            if not nm:
                continue
            lt = lt_default
            if i_lt is not None and r[i_lt]:
                lt = _app__norm_lt_value(self, r[i_lt])
            contact = (r[i_contact] or "") if i_contact is not None and i_contact < len(r) else ""
            contact = str(contact).strip()
            area = (r[i_area] or "") if i_area is not None else ""
            area = str(area).strip()
            pr = float(r[i_pr]) if i_pr is not None and r[i_pr] not in (None, "") else 0.0
            ir = None
            if lt == 'Regular':
                if i_ir is not None and r[i_ir] not in (None, ""):
                    ir = float(r[i_ir])
                else:
                    ir = 0.20
            dr = ""
            if i_dr is not None and r[i_dr]:
                dr = str(r[i_dr])[:10].strip()
                datetime.strptime(dr, "%Y-%m-%d")
            nu = ""
            if i_nu is not None and r[i_nu]:
                nu = str(r[i_nu])[:10].strip()
                datetime.strptime(nu, "%Y-%m-%d")
            psod = 0
            if i_ps is not None and r[i_ps] not in (None, ""):
                try:
                    psod = 1 if int(r[i_ps]) >= 1 else 0
                except Exception:
                    psod = 0

            exists = self.db.get_client_info(nm, loan_type=lt)
            if exists:
                ok = self.db.update_client(nm, new_name=nm, principal=pr, date_released=dr or None, area=area, interest_rate=ir, new_until=nu, loan_type=lt, pay_start_offset_days=psod, contact_number=contact)
                if ok:
                    updated += 1
            else:
                ok = self.db.add_client(nm, principal=pr, date_released=dr or None, area=area, interest_rate=ir, new_until=nu, loan_type=lt, pay_start_offset_days=psod, contact_number=contact)
                if ok:
                    imported += 1
        except Exception as e:
            failed += 1
            if len(failure_samples) < 5:
                failure_samples.append(f"Row {row_number}: {e}")
            try:
                _log_exc(f"import_clients_from_excel.row_{row_number}", e)
            except Exception:
                pass
            continue

    try:
        self.refresh_clients()
        self.refresh_reports()
        self.refresh_data_grid()
    except Exception as __spina_exc:
        _log_suppressed_once('excpass_0649', 'suppressed exception excpass_0649', __spina_exc)
        pass
    details = f"Imported: {imported}\nUpdated: {updated}\nFailed: {failed}"
    if failure_samples:
        details += "\n\nFirst errors:\n" + "\n".join(failure_samples)
    messagebox.showinfo("Import", details)

def _app_import_missing(self):
    try:
        n = self.db.import_missing_clients_from_transactions()
    except Exception as e:
        messagebox.showerror("Import", str(e))
        return
    try:
        self.refresh_clients()
    except Exception as __spina_exc:
        _log_suppressed_once('excpass_0650', 'suppressed exception excpass_0650', __spina_exc)
        pass
    messagebox.showinfo("Import", f"Added {n} missing client(s) from transactions.")

